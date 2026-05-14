from flask import Blueprint, render_template, session, redirect, url_for, request, flash
from db import get_connection
from security_helpers import (
    permissao_obrigatoria,
    usuario_tem_permissao,
)

rh_bp = Blueprint("rh", __name__)


# ------------------------------------------
# MENU RH
# ------------------------------------------
@rh_bp.route("/menu")
@permissao_obrigatoria("RH", "MENU")
def menu_rh():
    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    id_usuario = session["id_usuario"]
    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")
    tipo_global = str(session.get("tipo_global") or "").strip().lower()

    if tipo_global == "superusuario":
        permissoes = {
            "pode_avaliacoes": True,
            "pode_importar_abastecimentos": True,
            "pode_consultar_abastecimentos": True,
            "pode_configuracoes_rh": True,
            "pode_funcionarios": True,
            "pode_consultar_funcionarios": True,
            "pode_movimentacoes_funcionarios": True,
        }
    else:
        permissoes = {
            "pode_avaliacoes": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "AVALIACOES"
            ),
            "pode_importar_abastecimentos": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "IMPORTAR_ABASTECIMENTOS"
            ),
            "pode_consultar_abastecimentos": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "CONSULTAR_ABASTECIMENTOS"
            ),
            "pode_configuracoes_rh": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "CONFIGURACOES_RH"
            ),
            "pode_funcionarios": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "FUNCIONARIOS"
            ),
            "pode_consultar_funcionarios": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "CONSULTAR_FUNCIONARIOS"
            ),
            "pode_movimentacoes_funcionarios": usuario_tem_permissao(
                id_usuario, cod_empresa, "RH", "MOVIMENTACOES_FUNCIONARIOS"
            ),
        }

    return render_template(
        "menu_rh.html",
        cod_empresa=cod_empresa,
        nome_empresa=nome_empresa,
        url_voltar=url_for("sistema.selecionar_sistema"),
        texto_voltar="← Voltar",
        **permissoes
    )


# ------------------------------------------
# AVALIAÇÕES
# ------------------------------------------
@rh_bp.route("/avaliacoes")
@permissao_obrigatoria(
    "RH",
    "AVALIACOES",
    redirecionar_para="rh.menu_rh",
)
def avaliacoes():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    return render_template(
        "rh_avaliacoes.html",
        cod_empresa=session.get("cod_empresa", ""),
        nome_empresa=session.get("nome_empresa", ""),
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )


# ------------------------------------------
# IMPORTAR ABASTECIMENTOS
# ------------------------------------------
@rh_bp.route("/abastecimentos/importar", methods=["GET", "POST"])
@permissao_obrigatoria(
    "RH",
    "IMPORTAR_ABASTECIMENTOS",
    redirecionar_para="rh.menu_rh",
)
def importar_abastecimentos():
    from openpyxl import load_workbook
    import tempfile
    import os
    import re
    import unicodedata
    from datetime import datetime, date
    from psycopg2.extras import execute_batch

    def normalizar(txt):
        txt = str(txt or "").strip().lower()
        txt = unicodedata.normalize("NFKD", txt)
        txt = "".join(c for c in txt if not unicodedata.combining(c))
        return txt

    def numero(valor):
        if valor is None:
            return 0
        if isinstance(valor, (int, float)):
            return float(valor)

        txt = str(valor).strip()
        if not txt:
            return 0

        txt = txt.replace(".", "").replace(",", ".")

        try:
            return float(txt)
        except:
            return 0

    def data_valida(valor):
        if isinstance(valor, datetime):
            return valor.date()

        if isinstance(valor, date):
            return valor

        if isinstance(valor, str):
            txt = valor.strip()

            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(txt, fmt).date()
                except:
                    pass

        return None

    def extrair_numero_abastecimentos(txt):
        txt = str(txt or "")
        nums = re.findall(r"\d+", txt.replace(".", ""))

        if not nums:
            return 0

        return int(nums[-1])

    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")
    resumo = None

    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo.", "error")
            return redirect(url_for("rh.importar_abastecimentos"))

        caminho_tmp = None
        conn = get_connection()
        cur = conn.cursor()

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                caminho_tmp = tmp.name
                arquivo.save(caminho_tmp)

            wb = load_workbook(caminho_tmp, data_only=True)
            ws = wb.active

            linhas = []
            mapa = {}

            cod_filial_atual = None
            cache_filiais = {}
            filiais_nao_encontradas = set()

            funcionario_atual = None
            indices_funcionario_atual = []

            for row in ws.iter_rows(values_only=True):
                if not any(row):
                    continue

                col_a_original = str(row[0] or "").strip()
                col_b_original = str(row[1] or "").strip()

                col_a = normalizar(col_a_original)
                col_b = col_b_original.strip()

                # ------------------------------------------
                # QUEBRA DE FILIAL
                # ------------------------------------------
                if "filial" in col_a:
                    nome_filial_importacao = col_b_original.strip()

                    funcionario_atual = None
                    indices_funcionario_atual = []

                    if nome_filial_importacao in cache_filiais:
                        cod_filial_atual = cache_filiais[nome_filial_importacao]
                    else:
                        cur.execute("""
                            SELECT cod_filial
                            FROM filiais
                            WHERE cod_empresa = %s
                              AND (
                                    UPPER(TRIM(COALESCE(nome_filial_importacao, ''))) = UPPER(TRIM(%s))
                                 OR UPPER(TRIM(COALESCE(nome_filial, ''))) = UPPER(TRIM(%s))
                              )
                            LIMIT 1
                        """, (
                            cod_empresa,
                            nome_filial_importacao,
                            nome_filial_importacao,
                        ))

                        filial = cur.fetchone()
                        cod_filial_atual = filial[0] if filial else None

                        if cod_filial_atual is None:
                            filiais_nao_encontradas.add(nome_filial_importacao)

                        cache_filiais[nome_filial_importacao] = cod_filial_atual

                    continue

                # ------------------------------------------
                # LINHA DE TOTAL DO FUNCIONÁRIO
                # Exemplo: Número de Abastecimentos: 3490
                # ------------------------------------------
                if "numero de abastecimentos" in col_a:
                    qtd_abast = extrair_numero_abastecimentos(col_a_original)

                    if indices_funcionario_atual and qtd_abast > 0:
                        linhas[indices_funcionario_atual[0]]["numero_abastecimentos"] = qtd_abast

                    funcionario_atual = None
                    indices_funcionario_atual = []
                    continue

                textos = [normalizar(x) for x in row]
                linha_txt = " ".join(textos)

                # ------------------------------------------
                # DETECTAR CABEÇALHO
                # ------------------------------------------
                if "data" in linha_txt and "produto" in linha_txt and "funcion" in linha_txt:
                    mapa = {}

                    for idx, nome in enumerate(textos):
                        if "data" in nome:
                            mapa["data"] = idx
                        elif "produto" in nome:
                            mapa["produto"] = idx
                        elif "funcion" in nome:
                            mapa["funcionario"] = idx
                        elif "q" in nome and (
                            "abas" in nome
                            or "abast" in nome
                            or "litro" in nome
                            or "volume" in nome
                        ):
                            mapa["quantidade"] = idx
                        elif "medio" in nome or "preco" in nome:
                            mapa["preco"] = idx
                        elif "valor" in nome or "abastec" in nome:
                            mapa["valor"] = idx

                    continue

                if not mapa:
                    continue

                try:
                    data_raw = row[mapa.get("data")]
                    produto = str(row[mapa.get("produto")] or "").strip()
                    funcionario = str(row[mapa.get("funcionario")] or "").strip()
                except:
                    continue

                data_abast = data_valida(data_raw)

                if not data_abast or not produto or not funcionario:
                    continue

                quantidade = numero(row[mapa["quantidade"]]) if "quantidade" in mapa else 0
                preco = numero(row[mapa["preco"]]) if "preco" in mapa else 0
                valor = numero(row[mapa["valor"]]) if "valor" in mapa else 0

                if valor == 0 and quantidade > 0 and preco > 0:
                    valor = quantidade * preco

                chave_funcionario = (
                    cod_filial_atual,
                    funcionario.strip().upper()
                )

                if funcionario_atual != chave_funcionario:
                    funcionario_atual = chave_funcionario
                    indices_funcionario_atual = []

                linhas.append({
                    "cod_empresa": cod_empresa,
                    "cod_filial": cod_filial_atual,
                    "data_abastecimento": data_abast,
                    "produto": produto,
                    "funcionario": funcionario,
                    "quantidade": quantidade,
                    "preco_unitario": preco,
                    "valor_total": valor,
                    "arquivo_origem": arquivo.filename,
                    "numero_abastecimentos": 0,
                })

                indices_funcionario_atual.append(len(linhas) - 1)

            if not linhas:
                flash("Nenhum dado válido encontrado.", "error")
                return redirect(url_for("rh.importar_abastecimentos"))

            # Se alguma filial do arquivo não existir na empresa atual,
            # cancela antes de apagar ou inserir dados.
            if filiais_nao_encontradas:
                flash(
                    "Importação cancelada. Filiais não encontradas para esta empresa: "
                    + ", ".join(sorted(filiais_nao_encontradas)),
                    "error"
                )
                return redirect(url_for("rh.importar_abastecimentos"))

            meses_importados = sorted(set(
                (l["data_abastecimento"].year, l["data_abastecimento"].month)
                for l in linhas
            ))

            for ano, mes in meses_importados:
                cur.execute("""
                    DELETE FROM abastecimentos
                    WHERE cod_empresa = %s
                      AND EXTRACT(YEAR FROM data_abastecimento) = %s
                      AND EXTRACT(MONTH FROM data_abastecimento) = %s
                """, (cod_empresa, ano, mes))

            linhas_insert = [
                (
                    l["cod_empresa"],
                    l["cod_filial"],
                    l["data_abastecimento"],
                    l["produto"],
                    l["funcionario"],
                    l["quantidade"],
                    l["preco_unitario"],
                    l["valor_total"],
                    l["arquivo_origem"],
                    l["numero_abastecimentos"],
                )
                for l in linhas
            ]

            execute_batch(cur, """
                INSERT INTO abastecimentos (
                    cod_empresa,
                    cod_filial,
                    data_abastecimento,
                    produto,
                    funcionario,
                    quantidade,
                    preco_unitario,
                    valor_total,
                    arquivo_origem,
                    numero_abastecimentos
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, linhas_insert, page_size=500)

            conn.commit()

            resumo = {
                "arquivo": arquivo.filename,
                "linhas": len(linhas),
                "total_quantidade": float(sum(l["quantidade"] or 0 for l in linhas)),
                "total_valor": float(sum(l["valor_total"] or 0 for l in linhas)),
                "total_abastecimentos": int(sum(l["numero_abastecimentos"] or 0 for l in linhas)),
                "meses_importados": ", ".join([f"{mes:02d}/{ano}" for ano, mes in meses_importados]),
            }

            flash(f"{len(linhas)} registros importados com sucesso.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao importar: {e}", "error")

        finally:
            cur.close()
            conn.close()

            if caminho_tmp and os.path.exists(caminho_tmp):
                os.remove(caminho_tmp)

    return render_template(
        "importar_abastecimentos.html",
        cod_empresa=cod_empresa,
        nome_empresa=nome_empresa,
        resumo=resumo,
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )
# ------------------------------------------
# CONSULTAR ABASTECIMENTOS
# ------------------------------------------
@rh_bp.route("/abastecimentos/consultar")
@permissao_obrigatoria(
    "RH",
    "CONSULTAR_ABASTECIMENTOS",
    redirecionar_para="rh.menu_rh",
)
def consultar_abastecimentos():
    from psycopg2.extras import RealDictCursor
    from datetime import datetime
    from zoneinfo import ZoneInfo

    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    hoje = datetime.now(ZoneInfo("America/Sao_Paulo")).date()

    if hoje.month == 1:
        mes_padrao = 12
        ano_padrao = hoje.year - 1
    else:
        mes_padrao = hoje.month - 1
        ano_padrao = hoje.year

    ano_sel = (request.args.get("ano") or str(ano_padrao)).strip()
    mes_sel = (request.args.get("mes") or str(mes_padrao)).strip()
    filial_sel = (request.args.get("cod_filial") or "").strip()
    ordem_sel = (request.args.get("ordem") or "qtd").strip()

    ordens_sql = {
        "qtd": "qtd_abastecimentos DESC",
        "volume": "total_litros DESC",
        "valor": "total_valor DESC",
        "ticket": "ticket_medio DESC",
        "litros_abast": "litros_por_abastecimento DESC",
    }

    ordem_sql = ordens_sql.get(ordem_sel, "qtd_abastecimentos DESC")

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT DISTINCT
                a.cod_filial,
                COALESCE(f.nome_filial, 'Sem filial informada') AS nome_filial
            FROM abastecimentos a
            LEFT JOIN filiais f
              ON f.cod_empresa = a.cod_empresa
             AND f.cod_filial = a.cod_filial
            WHERE a.cod_empresa = %s
            ORDER BY a.cod_filial
        """, (cod_empresa,))
        filiais = cur.fetchall() or []

        filtros = ["a.cod_empresa = %s"]
        params = [cod_empresa]

        if ano_sel:
            filtros.append("EXTRACT(YEAR FROM a.data_abastecimento) = %s")
            params.append(int(ano_sel))

        if mes_sel:
            filtros.append("EXTRACT(MONTH FROM a.data_abastecimento) = %s")
            params.append(int(mes_sel))

        if filial_sel:
            filtros.append("a.cod_filial = %s")
            params.append(int(filial_sel))

        where_sql = " AND ".join(filtros)

        cur.execute(f"""
            SELECT
                COALESCE(a.cod_filial, 0) AS cod_filial,
                COALESCE(f.nome_filial, 'Sem filial informada') AS nome_filial,
                a.funcionario,
                COALESCE(SUM(a.numero_abastecimentos), 0) AS qtd_abastecimentos,
                COALESCE(SUM(a.quantidade), 0) AS total_litros,
                COALESCE(SUM(a.valor_total), 0) AS total_valor,
                CASE
                    WHEN COALESCE(SUM(a.numero_abastecimentos), 0) > 0
                    THEN COALESCE(SUM(a.valor_total), 0) / COALESCE(SUM(a.numero_abastecimentos), 0)
                    ELSE 0
                END AS ticket_medio,
                CASE
                    WHEN COALESCE(SUM(a.numero_abastecimentos), 0) > 0
                    THEN COALESCE(SUM(a.quantidade), 0) / COALESCE(SUM(a.numero_abastecimentos), 0)
                    ELSE 0
                END AS litros_por_abastecimento
            FROM abastecimentos a
            LEFT JOIN filiais f
              ON f.cod_empresa = a.cod_empresa
             AND f.cod_filial = a.cod_filial
            WHERE {where_sql}
            GROUP BY
                a.cod_filial,
                f.nome_filial,
                a.funcionario
            ORDER BY
                COALESCE(a.cod_filial, 0),
                {ordem_sql},
                a.funcionario
        """, params)

        linhas = cur.fetchall() or []

        heatmap = {}

        for l in linhas:
            cod = l["cod_filial"] or 0

            if cod not in heatmap:
                heatmap[cod] = {
                    "qtd": {"min": l["qtd_abastecimentos"] or 0, "max": l["qtd_abastecimentos"] or 0},
                    "litros": {"min": l["total_litros"] or 0, "max": l["total_litros"] or 0},
                    "valor": {"min": l["total_valor"] or 0, "max": l["total_valor"] or 0},
                    "ticket": {"min": l["ticket_medio"] or 0, "max": l["ticket_medio"] or 0},
                    "litros_abast": {"min": l["litros_por_abastecimento"] or 0, "max": l["litros_por_abastecimento"] or 0},
                }
            else:
                heatmap[cod]["qtd"]["min"] = min(heatmap[cod]["qtd"]["min"], l["qtd_abastecimentos"] or 0)
                heatmap[cod]["qtd"]["max"] = max(heatmap[cod]["qtd"]["max"], l["qtd_abastecimentos"] or 0)

                heatmap[cod]["litros"]["min"] = min(heatmap[cod]["litros"]["min"], l["total_litros"] or 0)
                heatmap[cod]["litros"]["max"] = max(heatmap[cod]["litros"]["max"], l["total_litros"] or 0)

                heatmap[cod]["valor"]["min"] = min(heatmap[cod]["valor"]["min"], l["total_valor"] or 0)
                heatmap[cod]["valor"]["max"] = max(heatmap[cod]["valor"]["max"], l["total_valor"] or 0)

                heatmap[cod]["ticket"]["min"] = min(heatmap[cod]["ticket"]["min"], l["ticket_medio"] or 0)
                heatmap[cod]["ticket"]["max"] = max(heatmap[cod]["ticket"]["max"], l["ticket_medio"] or 0)

                heatmap[cod]["litros_abast"]["min"] = min(heatmap[cod]["litros_abast"]["min"], l["litros_por_abastecimento"] or 0)
                heatmap[cod]["litros_abast"]["max"] = max(heatmap[cod]["litros_abast"]["max"], l["litros_por_abastecimento"] or 0)

        totais_filiais = {}

        for l in linhas:
            cod = l["cod_filial"] or 0

            if cod not in totais_filiais:
                totais_filiais[cod] = {
                    "qtd_abastecimentos": 0,
                    "total_litros": 0,
                    "total_valor": 0,
                    "ticket_medio": 0,
                    "litros_por_abastecimento": 0,
                }

            totais_filiais[cod]["qtd_abastecimentos"] += l["qtd_abastecimentos"] or 0
            totais_filiais[cod]["total_litros"] += l["total_litros"] or 0
            totais_filiais[cod]["total_valor"] += l["total_valor"] or 0

        for cod, t in totais_filiais.items():
            qtd = t["qtd_abastecimentos"] or 0
            if qtd > 0:
                t["ticket_medio"] = t["total_valor"] / qtd
                t["litros_por_abastecimento"] = t["total_litros"] / qtd

        cur.execute(f"""
            SELECT
                COALESCE(SUM(a.numero_abastecimentos), 0) AS qtd_abastecimentos,
                COALESCE(SUM(a.quantidade), 0) AS total_litros,
                COALESCE(SUM(a.valor_total), 0) AS total_valor,
                CASE
                    WHEN COALESCE(SUM(a.numero_abastecimentos), 0) > 0
                    THEN COALESCE(SUM(a.valor_total), 0) / COALESCE(SUM(a.numero_abastecimentos), 0)
                    ELSE 0
                END AS ticket_medio,
                CASE
                    WHEN COALESCE(SUM(a.numero_abastecimentos), 0) > 0
                    THEN COALESCE(SUM(a.quantidade), 0) / COALESCE(SUM(a.numero_abastecimentos), 0)
                    ELSE 0
                END AS litros_por_abastecimento
            FROM abastecimentos a
            WHERE {where_sql}
        """, params)

        totais = cur.fetchone() or {}

    finally:
        cur.close()
        conn.close()

    return render_template(
        "consultar_abastecimentos.html",
        cod_empresa=cod_empresa,
        nome_empresa=nome_empresa,
        ano_sel=ano_sel,
        mes_sel=mes_sel,
        filial_sel=filial_sel,
        ordem_sel=ordem_sel,
        filiais=filiais,
        linhas=linhas,
        totais=totais,
        heatmap=heatmap,
        totais_filiais=totais_filiais,
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )

# ------------------------------------------
# CONFIGURAÇÕES RH
# ------------------------------------------
@rh_bp.route("/configuracoes")
@permissao_obrigatoria(
    "RH",
    "CONFIGURACOES_RH",
    redirecionar_para="rh.menu_rh",
)
def configuracoes_rh():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    return render_template(
        "menu_configuracoes_rh.html",
        cod_empresa=session.get("cod_empresa", ""),
        nome_empresa=session.get("nome_empresa", ""),
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )

# ------------------------------------------
# FUNCIONÁRIOS
# ------------------------------------------
@rh_bp.route("/funcionarios", methods=["GET", "POST"])
@permissao_obrigatoria(
    "RH",
    "FUNCIONARIOS",
    redirecionar_para="rh.menu_rh",
)
def funcionarios():
    from psycopg2.extras import RealDictCursor

    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if request.method == "POST":
            matricula = (request.form.get("matricula") or "").strip()
            nome = (request.form.get("nome") or "").strip()
            email = (request.form.get("email") or "").strip()

            if not nome:
                flash("Informe o nome do funcionário.", "error")
                return redirect(url_for("rh.funcionarios"))

            cur.execute("""
                INSERT INTO funcionarios (
                    cod_empresa,
                    matricula,
                    nome,
                    email,
                    ativo
                )
                VALUES (%s, NULLIF(%s, ''), %s, NULLIF(%s, ''), TRUE)
            """, (
                cod_empresa,
                matricula,
                nome,
                email,
            ))

            conn.commit()
            flash("Funcionário cadastrado com sucesso.", "success")
            return redirect(url_for("rh.funcionarios"))

        cur.execute("""
            SELECT
                f.id,
                f.matricula,
                f.nome,
                f.email,
                f.cod_filial,
                fi.nome_filial,
                f.id_cargo,
                c.descricao AS cargo,
                f.data_admissao,
                f.data_demissao,
                f.ativo
            FROM funcionarios f
            LEFT JOIN filiais fi
              ON fi.cod_empresa = f.cod_empresa
             AND fi.cod_filial = f.cod_filial
            LEFT JOIN cargos c
              ON c.id = f.id_cargo
             AND c.cod_empresa = f.cod_empresa
            WHERE f.cod_empresa = %s
            ORDER BY
                f.ativo DESC,
                f.nome
        """, (cod_empresa,))

        funcionarios_lista = cur.fetchall() or []

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao processar funcionários: {e}", "error")
        funcionarios_lista = []

    finally:
        cur.close()
        conn.close()

    return render_template(
        "funcionarios.html",
        cod_empresa=cod_empresa,
        nome_empresa=nome_empresa,
        funcionarios=funcionarios_lista,
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )


# ------------------------------------------
# EDITAR FUNCIONÁRIO
# Somente dados cadastrais.
# Filial, cargo, admissão, demissão e ativo
# são controlados pela tela de movimentações.
# ------------------------------------------
@rh_bp.route("/funcionarios/<int:id_funcionario>/editar", methods=["POST"])
@permissao_obrigatoria(
    "RH",
    "FUNCIONARIOS",
    redirecionar_para="rh.menu_rh",
)
def editar_funcionario(id_funcionario):
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()

    matricula = (request.form.get("matricula") or "").strip()
    nome = (request.form.get("nome") or "").strip()
    email = (request.form.get("email") or "").strip()

    if not nome:
        flash("Informe o nome do funcionário.", "error")
        return redirect(url_for("rh.funcionarios"))

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE funcionarios
               SET matricula = NULLIF(%s, ''),
                   nome = %s,
                   email = NULLIF(%s, ''),
                   atualizado_em = NOW()
             WHERE id = %s
               AND cod_empresa = %s
        """, (
            matricula,
            nome,
            email,
            id_funcionario,
            cod_empresa,
        ))

        conn.commit()
        flash("Funcionário atualizado com sucesso.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao atualizar funcionário: {e}", "error")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("rh.funcionarios"))


# ------------------------------------------
# MOVIMENTAÇÕES DE FUNCIONÁRIOS
# ------------------------------------------
@rh_bp.route("/funcionarios/movimentacoes", methods=["GET", "POST"])
@permissao_obrigatoria(
    "RH",
    "MOVIMENTACOES_FUNCIONARIOS",
    redirecionar_para="rh.menu_rh",
)
def movimentacoes_funcionarios():
    from psycopg2.extras import RealDictCursor

    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        if request.method == "POST":
            id_funcionario = request.form.get("id_funcionario")
            tipo_movimento = (request.form.get("tipo_movimento") or "").strip().upper()
            data_movimento = request.form.get("data_movimento")
            cod_filial_destino = request.form.get("cod_filial_destino") or None
            id_cargo_novo = request.form.get("id_cargo_novo") or None
            observacao = (request.form.get("observacao") or "").strip()

            if not id_funcionario or not tipo_movimento or not data_movimento:
                flash("Informe funcionário, tipo de movimento e data.", "error")
                return redirect(url_for("rh.movimentacoes_funcionarios"))

            cur.execute("""
                SELECT
                    id,
                    cod_filial,
                    id_cargo,
                    data_admissao,
                    data_demissao,
                    ativo
                FROM funcionarios
                WHERE id = %s
                  AND cod_empresa = %s
            """, (id_funcionario, cod_empresa))

            funcionario = cur.fetchone()

            if not funcionario:
                flash("Funcionário não encontrado.", "error")
                return redirect(url_for("rh.movimentacoes_funcionarios"))

            cod_filial_origem = funcionario["cod_filial"]
            id_cargo_anterior = funcionario["id_cargo"]

            if tipo_movimento == "ADMISSAO":
                if not cod_filial_destino or not id_cargo_novo:
                    flash("Na admissão, informe filial e cargo.", "error")
                    return redirect(url_for("rh.movimentacoes_funcionarios"))

                cur.execute("""
                    INSERT INTO funcionarios_movimentacoes (
                        cod_empresa,
                        id_funcionario,
                        tipo_movimento,
                        data_movimento,
                        cod_filial_origem,
                        cod_filial_destino,
                        id_cargo_anterior,
                        id_cargo_novo,
                        observacao
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    cod_empresa,
                    id_funcionario,
                    tipo_movimento,
                    data_movimento,
                    cod_filial_origem,
                    cod_filial_destino,
                    id_cargo_anterior,
                    id_cargo_novo,
                    observacao,
                ))

                cur.execute("""
                    UPDATE funcionarios
                       SET cod_filial = %s,
                           id_cargo = %s,
                           data_admissao = %s,
                           data_demissao = NULL,
                           ativo = TRUE,
                           atualizado_em = NOW()
                     WHERE id = %s
                       AND cod_empresa = %s
                """, (
                    cod_filial_destino,
                    id_cargo_novo,
                    data_movimento,
                    id_funcionario,
                    cod_empresa,
                ))

            elif tipo_movimento == "TRANSFERENCIA":
                if not cod_filial_destino:
                    flash("Na transferência, informe a filial de destino.", "error")
                    return redirect(url_for("rh.movimentacoes_funcionarios"))

                cargo_final = id_cargo_novo or id_cargo_anterior

                cur.execute("""
                    INSERT INTO funcionarios_movimentacoes (
                        cod_empresa,
                        id_funcionario,
                        tipo_movimento,
                        data_movimento,
                        cod_filial_origem,
                        cod_filial_destino,
                        id_cargo_anterior,
                        id_cargo_novo,
                        observacao
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    cod_empresa,
                    id_funcionario,
                    tipo_movimento,
                    data_movimento,
                    cod_filial_origem,
                    cod_filial_destino,
                    id_cargo_anterior,
                    cargo_final,
                    observacao,
                ))

                cur.execute("""
                    UPDATE funcionarios
                       SET cod_filial = %s,
                           id_cargo = %s,
                           ativo = TRUE,
                           atualizado_em = NOW()
                     WHERE id = %s
                       AND cod_empresa = %s
                """, (
                    cod_filial_destino,
                    cargo_final,
                    id_funcionario,
                    cod_empresa,
                ))

            elif tipo_movimento == "ALTERACAO_CARGO":
                if not id_cargo_novo:
                    flash("Na alteração de cargo, informe o novo cargo.", "error")
                    return redirect(url_for("rh.movimentacoes_funcionarios"))

                cur.execute("""
                    INSERT INTO funcionarios_movimentacoes (
                        cod_empresa,
                        id_funcionario,
                        tipo_movimento,
                        data_movimento,
                        cod_filial_origem,
                        cod_filial_destino,
                        id_cargo_anterior,
                        id_cargo_novo,
                        observacao
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    cod_empresa,
                    id_funcionario,
                    tipo_movimento,
                    data_movimento,
                    cod_filial_origem,
                    cod_filial_origem,
                    id_cargo_anterior,
                    id_cargo_novo,
                    observacao,
                ))

                cur.execute("""
                    UPDATE funcionarios
                       SET id_cargo = %s,
                           atualizado_em = NOW()
                     WHERE id = %s
                       AND cod_empresa = %s
                """, (
                    id_cargo_novo,
                    id_funcionario,
                    cod_empresa,
                ))

            elif tipo_movimento == "DEMISSAO":
                cur.execute("""
                    INSERT INTO funcionarios_movimentacoes (
                        cod_empresa,
                        id_funcionario,
                        tipo_movimento,
                        data_movimento,
                        cod_filial_origem,
                        cod_filial_destino,
                        id_cargo_anterior,
                        id_cargo_novo,
                        observacao
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    cod_empresa,
                    id_funcionario,
                    tipo_movimento,
                    data_movimento,
                    cod_filial_origem,
                    None,
                    id_cargo_anterior,
                    None,
                    observacao,
                ))

                cur.execute("""
                    UPDATE funcionarios
                       SET data_demissao = %s,
                           ativo = FALSE,
                           atualizado_em = NOW()
                     WHERE id = %s
                       AND cod_empresa = %s
                """, (
                    data_movimento,
                    id_funcionario,
                    cod_empresa,
                ))

            else:
                flash("Tipo de movimento inválido.", "error")
                return redirect(url_for("rh.movimentacoes_funcionarios"))

            conn.commit()
            flash("Movimentação registrada com sucesso.", "success")
            return redirect(url_for("rh.movimentacoes_funcionarios"))

        cur.execute("""
            SELECT
                f.id,
                f.matricula,
                f.nome,
                f.email,
                f.cod_filial,
                fi.nome_filial,
                f.id_cargo,
                c.descricao AS cargo,
                f.data_admissao,
                f.data_demissao,
                f.ativo
            FROM funcionarios f
            LEFT JOIN filiais fi
              ON fi.cod_empresa = f.cod_empresa
             AND fi.cod_filial = f.cod_filial
            LEFT JOIN cargos c
              ON c.id = f.id_cargo
             AND c.cod_empresa = f.cod_empresa
            WHERE f.cod_empresa = %s
            ORDER BY f.nome
        """, (cod_empresa,))
        funcionarios_lista = cur.fetchall() or []

        cur.execute("""
            SELECT
                cod_filial,
                nome_filial
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY nome_filial
        """, (cod_empresa,))
        filiais = cur.fetchall() or []

        cur.execute("""
            SELECT
                id,
                codigo,
                descricao
            FROM cargos
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY descricao
        """, (cod_empresa,))
        cargos = cur.fetchall() or []

        cur.execute("""
            SELECT
                m.id,
                m.tipo_movimento,
                m.data_movimento,
                m.observacao,
                f.nome AS funcionario,
                fo.nome_filial AS filial_origem,
                fd.nome_filial AS filial_destino,
                ca.descricao AS cargo_anterior,
                cn.descricao AS cargo_novo
            FROM funcionarios_movimentacoes m
            INNER JOIN funcionarios f
               ON f.id = m.id_funcionario
              AND f.cod_empresa = m.cod_empresa
            LEFT JOIN filiais fo
              ON fo.cod_empresa = m.cod_empresa
             AND fo.cod_filial = m.cod_filial_origem
            LEFT JOIN filiais fd
              ON fd.cod_empresa = m.cod_empresa
             AND fd.cod_filial = m.cod_filial_destino
            LEFT JOIN cargos ca
              ON ca.id = m.id_cargo_anterior
             AND ca.cod_empresa = m.cod_empresa
            LEFT JOIN cargos cn
              ON cn.id = m.id_cargo_novo
             AND cn.cod_empresa = m.cod_empresa
            WHERE m.cod_empresa = %s
            ORDER BY
                m.data_movimento DESC,
                m.id DESC
            LIMIT 100
        """, (cod_empresa,))
        movimentacoes = cur.fetchall() or []

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao processar movimentações: {e}", "error")
        funcionarios_lista = []
        filiais = []
        cargos = []
        movimentacoes = []

    finally:
        cur.close()
        conn.close()

    return render_template(
        "movimentacoes_funcionarios.html",
        cod_empresa=cod_empresa,
        nome_empresa=nome_empresa,
        funcionarios=funcionarios_lista,
        filiais=filiais,
        cargos=cargos,
        movimentacoes=movimentacoes,
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )

# ------------------------------------------
# CONSULTAR FUNCIONÁRIOS
# ------------------------------------------
@rh_bp.route("/funcionarios/consultar")
@permissao_obrigatoria(
    "RH",
    "CONSULTAR_FUNCIONARIOS",
    redirecionar_para="rh.menu_rh",
)
def consultar_funcionarios():
    from psycopg2.extras import RealDictCursor

    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    filial_sel = (request.args.get("cod_filial") or "").strip()
    somente_ativos = (request.args.get("somente_ativos") or "S").strip().upper()

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                cod_filial,
                nome_filial
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY cod_filial
        """, (cod_empresa,))
        filiais = cur.fetchall() or []

        filtros = ["f.cod_empresa = %s"]
        params = [cod_empresa]

        if filial_sel:
            filtros.append("f.cod_filial = %s")
            params.append(int(filial_sel))

        if somente_ativos == "S":
            filtros.append("f.ativo = TRUE")

        where_sql = " AND ".join(filtros)

        cur.execute(f"""
            SELECT
                f.id,
                f.matricula,
                f.nome,
                f.email,
                f.cod_filial,
                COALESCE(fi.nome_filial, 'Sem filial informada') AS nome_filial,
                f.id_cargo,
                COALESCE(c.descricao, 'Sem cargo informado') AS cargo,
                f.data_admissao,
                f.data_demissao,
                f.ativo
            FROM funcionarios f
            LEFT JOIN filiais fi
              ON fi.cod_empresa = f.cod_empresa
             AND fi.cod_filial = f.cod_filial
            LEFT JOIN cargos c
              ON c.id = f.id_cargo
             AND c.cod_empresa = f.cod_empresa
            WHERE {where_sql}
            ORDER BY
                COALESCE(f.cod_filial, 999999),
                COALESCE(f.id_cargo, 999999),
                f.nome
        """, params)

        funcionarios = cur.fetchall() or []

        totais_filial = {}

        for f in funcionarios:
            cod_filial = f["cod_filial"] if f["cod_filial"] is not None else 0
            nome_filial = f["nome_filial"] or "Sem filial informada"
            cargo = f["cargo"] or "Sem cargo informado"

            if cod_filial not in totais_filial:
                totais_filial[cod_filial] = {
                    "nome_filial": nome_filial,
                    "total": 0,
                    "cargos": {}
                }

            totais_filial[cod_filial]["total"] += 1
            totais_filial[cod_filial]["cargos"][cargo] = (
                totais_filial[cod_filial]["cargos"].get(cargo, 0) + 1
            )

    except Exception as e:
        flash(f"Erro ao consultar funcionários: {e}", "error")
        filiais = []
        funcionarios = []
        totais_filial = {}

    finally:
        cur.close()
        conn.close()

    return render_template(
        "consultar_funcionarios.html",
        cod_empresa=cod_empresa,
        nome_empresa=nome_empresa,
        filiais=filiais,
        funcionarios=funcionarios,
        totais_filial=totais_filial,
        filial_sel=filial_sel,
        somente_ativos=somente_ativos,
        url_voltar=url_for("rh.menu_rh"),
        texto_voltar="← Voltar",
    )