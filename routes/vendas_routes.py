from flask import Blueprint, render_template, session, redirect, url_for, request, flash, g
from security_helpers import permissao_obrigatoria
import psycopg2
from psycopg2.extras import RealDictCursor, execute_values
from openpyxl import load_workbook
import tempfile
import os
import re
import uuid
import calendar
import csv
import io

from collections import defaultdict
from datetime import date, timedelta, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from zoneinfo import ZoneInfo

vendas_bp = Blueprint("vendas", __name__, url_prefix="/vendas")


# =========================
# BANCO
# =========================
def get_connection():
    return psycopg2.connect(
        dbname="postgres",
        user="postgres.uaafkuovkzkozmscyapw",
        password="DataMatrix@1962#",
        host="aws-1-us-east-1.pooler.supabase.com",
        port=6543,
        sslmode="require"
    )


# =========================
# EXTRAIR PERIODO DO ARQUIVO
# =========================



def extrair_periodo_do_arquivo_diario(ws):
    datas = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        linha = list(row or [])

        while len(linha) < 2:
            linha.append(None)

        col_b = str(linha[1] or "").strip()

        # Exemplo: 01/04/2026 - Quarta
        m = re.match(r"^(\d{2}/\d{2}/\d{4})\s*-\s*.+$", col_b)
        if m:
            try:
                data_lida = datetime.strptime(m.group(1), "%d/%m/%Y").date()
                datas.append(data_lida)
            except Exception:
                pass

    if not datas:
        return None, None

    return min(datas), max(datas)
# =========================
# FUNÇÕES AUXILIARES
# =========================
def para_float(valor):
    if valor is None or valor == "":
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    try:
        return float(texto)
    except Exception:
        pass

    texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0


def para_decimal(valor):
    if valor is None or valor == "":
        return Decimal("0")

    if isinstance(valor, Decimal):
        return valor

    if isinstance(valor, (int, float)):
        return Decimal(str(valor))

    texto = str(valor).strip()

    try:
        return Decimal(texto)
    except Exception:
        pass

    texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def para_data_excel(valor):
    if valor is None or valor == "":
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()

    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt).date()
        except Exception:
            pass

    return None


def formatar_numero_br(valor, casas=2):
    try:
        numero = float(valor or 0)
    except Exception:
        numero = 0.0

    texto = f"{numero:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_data_brasil(dt):
    if not dt:
        return ""

    try:
        tz_brasil = ZoneInfo("America/Fortaleza")

        if hasattr(dt, "tzinfo") and dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))

        if hasattr(dt, "astimezone"):
            dt_local = dt.astimezone(tz_brasil)
            return dt_local.strftime("%d/%m/%y")

        return dt.strftime("%d/%m/%y")
    except Exception:
        return dt.strftime("%d/%m/%y")

def normalizar_nome_filial_importacao(nome):
    if not nome:
        return ""

    texto = str(nome).strip().upper()
    texto = re.sub(r"\s+", " ", texto)

    return texto


def localizar_filial_por_nome_importacao(nome_planilha, mapa_filiais):
    nome_planilha_norm = normalizar_nome_filial_importacao(nome_planilha)

    if not nome_planilha_norm:
        return None

    for nome_padrao, cod_filial in mapa_filiais.items():
        tamanho = len(nome_padrao)

        if nome_planilha_norm[:tamanho] == nome_padrao:
            return cod_filial

    return None


def cor_excel_51(valor, minimo, maximo):
    try:
        v = float(valor)
        mn = float(minimo)
        mx = float(maximo)
    except Exception:
        return "#ffffff"

    if mx == mn:
        return "#f5da90"

    if mx < mn:
        return "#ffffff"

    ratio = (v - mn) / (mx - mn)
    ratio = max(0, min(1, ratio))

    faixa = int(round(ratio * 50))
    faixa = max(0, min(50, faixa))

    cores = [
        "#f8696b", "#f96d6c", "#f9716d", "#fa756e", "#fa796f",
        "#fb7d70", "#fb8171", "#fc8572", "#fc8973", "#fd8d74",
        "#fd9175", "#fe9576", "#fe9977", "#ef9e78", "#f0a37a",
        "#f0a87c", "#f1ad7e", "#f1b280", "#f2b782", "#f2bc84",
        "#f3c186", "#f3c688", "#f4cb8a", "#f4d08c", "#f5d58e",
        "#f5da90", "#efe08f", "#e9e58e", "#e3ea8d", "#dde68d",
        "#d7e28c", "#d1df8b", "#cbdb8a", "#c5d789", "#bfd489",
        "#b9d088", "#b3cc87", "#add986", "#a7c585", "#a1c184",
        "#9bbe84", "#95ba83", "#8fb682", "#89b281", "#83af80",
        "#7dab80", "#77a77f", "#71a37e", "#6ba07d", "#67c07b",
        "#63be7b"
    ]
    return cores[faixa]


def aplicar_heatmap_na_grade(grade):
    valores = []

    for linha in grade["linhas"]:
        for v in linha["valores"]:
            if v not in (None, 0, 0.0, ""):
                valores.append(float(v))

    if not valores:
        grade["cores_linhas"] = []
        return grade

    minimo = min(valores)
    maximo = max(valores)

    cores_linhas = []
    for linha in grade["linhas"]:
        cores = []
        for v in linha["valores"]:
            if v in (None, 0, 0.0, ""):
                cores.append("")
            else:
                cores.append(cor_excel_51(v, minimo, maximo))
        cores_linhas.append(cores)

    grade["cores_linhas"] = cores_linhas
    return grade


def aplicar_heatmap_na_coluna_total(grade):
    valores = []

    for linha in grade["linhas"]:
        v = linha.get("total")
        if v not in (None, 0, 0.0, ""):
            valores.append(float(v))

    if not valores:
        grade["cores_totais"] = []
        return grade

    minimo = min(valores)
    maximo = max(valores)

    cores_totais = []
    for linha in grade["linhas"]:
        v = linha.get("total")
        if v in (None, 0, 0.0, ""):
            cores_totais.append("")
        else:
            cores_totais.append(cor_excel_51(v, minimo, maximo))

    grade["cores_totais"] = cores_totais
    return grade


def aplicar_heatmap_variacoes(linhas, campos):
    if not linhas:
        return linhas

    for campo in campos:
        valores = []
        for linha in linhas:
            valor = linha.get(campo)
            if valor not in (None, 0, 0.0, ""):
                valores.append(float(valor))

        minimo = min(valores) if valores else None
        maximo = max(valores) if valores else None

        cor_key = f"{campo}_cor"

        for linha in linhas:
            valor = linha.get(campo)
            if minimo is None or valor in (None, 0, 0.0, ""):
                linha[cor_key] = ""
            else:
                linha[cor_key] = cor_excel_51(valor, minimo, maximo)

    return linhas


def aplicar_heatmap_consulta(linhas):
    if not linhas:
        return linhas
    return aplicar_heatmap_variacoes(linhas, ["quantidade", "valor", "mb", "mun"])



def serie_grafico_consulta(linhas):
    """Série enxuta para os gráficos da tela de consultas (nada persistido)."""
    return [
        {
            "data": linha.get("data_fmt"),
            "quantidade": float(linha.get("quantidade") or 0),
            "valor": float(linha.get("valor") or 0),
            "mb": float(linha.get("mb") or 0),
            "mun": float(linha.get("mun") or 0),
        }
        for linha in (linhas or [])
    ]


def localizar_total_filial(ws, nome_busca):
    nome_busca = (nome_busca or "").strip().upper()
    if not nome_busca:
        return None

    for linha in range(1, 2000):
        valor = ws.cell(row=linha, column=1).value
        texto = str(valor).strip().upper() if valor is not None else ""

        if nome_busca in texto:
            for linha2 in range(linha, min(linha + 100, 2000)):
                valor2 = ws.cell(row=linha2, column=1).value
                texto2 = str(valor2).strip().upper() if valor2 is not None else ""

                if texto2 == "TOTAL FILIAL:":
                    return linha2

    return None


def upsert(cur, tabela, campo, cod_empresa, cod_filial, ano, mes, valor):
    cur.execute(
        f"""
        INSERT INTO {tabela}
        (cod_empresa, cod_filial, ano, mes, data_importacao, {campo})
        VALUES (%s, %s, %s, %s, NOW(), %s)
        ON CONFLICT (cod_empresa, cod_filial, ano, mes)
        DO UPDATE SET
            {campo} = EXCLUDED.{campo},
            data_importacao = NOW()
        """,
        (cod_empresa, cod_filial, ano, mes, valor)
    )


def limpar_importacao_mes(cur, cod_empresa, ano, mes):
    cur.execute("""
        DELETE FROM vendas_unidades_sintetico
        WHERE cod_empresa = %s
          AND ano = %s
          AND mes = %s
    """, (cod_empresa, ano, mes))

    cur.execute("""
        DELETE FROM vendas_valores_sintetico
        WHERE cod_empresa = %s
          AND ano = %s
          AND mes = %s
    """, (cod_empresa, ano, mes))

    cur.execute("""
        DELETE FROM vendas_mb_sintetico
        WHERE cod_empresa = %s
          AND ano = %s
          AND mes = %s
    """, (cod_empresa, ano, mes))


def limpar_importacao_diaria_periodo(cur, conn, cod_empresa, data_ini, data_fim, lote=5000):
    while True:
        cur.execute("""
            DELETE FROM vendas_diarias
            WHERE ctid IN (
                SELECT ctid
                FROM vendas_diarias
                WHERE cod_empresa = %s
                  AND data BETWEEN %s AND %s
                LIMIT %s
            )
        """, (cod_empresa, data_ini, data_fim, lote))

        if cur.rowcount == 0:
            break

        conn.commit()




def inserir_importacao_painel(
    cur,
    cod_empresa,
    ano,
    mes,
    dia_base,
    dias_mes,
    quantidade_proj,
    valor_proj,
    mb_proj
):
    data_brasil = datetime.now(ZoneInfo("America/Sao_Paulo")).date()

    cur.execute("""
        INSERT INTO vendas_painel_importacoes
        (
            cod_empresa,
            ano,
            mes,
            dia_base,
            dias_mes,
            quantidade_proj,
            valor_proj,
            mb_proj,
            data_importacao,
            data_importacao_dia
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), %s)
        ON CONFLICT (cod_empresa, data_importacao_dia)
        DO UPDATE SET
            ano = EXCLUDED.ano,
            mes = EXCLUDED.mes,
            dia_base = EXCLUDED.dia_base,
            dias_mes = EXCLUDED.dias_mes,
            quantidade_proj = EXCLUDED.quantidade_proj,
            valor_proj = EXCLUDED.valor_proj,
            mb_proj = EXCLUDED.mb_proj,
            data_importacao = NOW()
    """, (
        cod_empresa,
        ano,
        mes,
        dia_base,
        dias_mes,
        float(quantidade_proj),
        float(valor_proj),
        float(mb_proj),
        data_brasil
    ))


# =========================
# PARÂMETROS DE VENDAS (por empresa)
# =========================
# Origem da importação do painel. A chave é o que fica gravado em
# vendas_parametros.sistema_origem_painel; o valor é o rótulo da tela.
SISTEMAS_ORIGEM_PAINEL = {
    "WEBPOSTOS": "WebPostos",
    "OCLOSET": "Sistema O Closet",
}

SISTEMA_ORIGEM_PAINEL_PADRAO = "WEBPOSTOS"

# Mesma ideia para a importação de VENDAS DIÁRIAS, em coluna própria
# (vendas_parametros.sistema_origem_diarias) — as duas importações são
# independentes, uma empresa pode ter painel de um jeito e diárias de outro.
SISTEMAS_ORIGEM_DIARIAS = {
    "WEBPOSTOS": "WebPostos",
    "OCLOSET": "Sistema O Closet",
}

SISTEMA_ORIGEM_DIARIAS_PADRAO = "WEBPOSTOS"


def _obter_origem_vendas(cod_empresa, coluna, opcoes, padrao):
    """Empresa sem linha em vendas_parametros cai no padrão (WebPostos)."""
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            f"""
            SELECT {coluna}
            FROM vendas_parametros
            WHERE cod_empresa = %s
            """,
            (cod_empresa,)
        )
        linha = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if not linha:
        return padrao

    sistema = str(linha[0] or "").strip().upper()
    if sistema not in opcoes:
        return padrao

    return sistema


def obter_sistema_origem_painel(cod_empresa):
    return _obter_origem_vendas(
        cod_empresa, "sistema_origem_painel",
        SISTEMAS_ORIGEM_PAINEL, SISTEMA_ORIGEM_PAINEL_PADRAO
    )


def obter_sistema_origem_diarias(cod_empresa):
    return _obter_origem_vendas(
        cod_empresa, "sistema_origem_diarias",
        SISTEMAS_ORIGEM_DIARIAS, SISTEMA_ORIGEM_DIARIAS_PADRAO
    )


@vendas_bp.app_template_global("origem_painel_empresa")
def origem_painel_empresa():
    """Origem do painel da empresa da sessão, para os templates.

    Cacheada em `g` porque a tela de importação chama mais de uma vez por
    requisição e a consulta é sempre a mesma.
    """
    if "cod_empresa" not in session:
        return {"sistema": SISTEMA_ORIGEM_PAINEL_PADRAO,
                "rotulo": SISTEMAS_ORIGEM_PAINEL[SISTEMA_ORIGEM_PAINEL_PADRAO]}

    if not hasattr(g, "_origem_painel"):
        sistema = obter_sistema_origem_painel(str(session["cod_empresa"]).strip())
        g._origem_painel = {
            "sistema": sistema,
            "rotulo": SISTEMAS_ORIGEM_PAINEL[sistema],
        }

    return g._origem_painel


@vendas_bp.app_template_global("origem_diarias_empresa")
def origem_diarias_empresa():
    """Origem das vendas diárias da empresa da sessão, para os templates."""
    if "cod_empresa" not in session:
        return {"sistema": SISTEMA_ORIGEM_DIARIAS_PADRAO,
                "rotulo": SISTEMAS_ORIGEM_DIARIAS[SISTEMA_ORIGEM_DIARIAS_PADRAO]}

    if not hasattr(g, "_origem_diarias"):
        sistema = obter_sistema_origem_diarias(str(session["cod_empresa"]).strip())
        g._origem_diarias = {
            "sistema": sistema,
            "rotulo": SISTEMAS_ORIGEM_DIARIAS[sistema],
        }

    return g._origem_diarias


def _gravar_origem_vendas(cod_empresa, coluna, sistema):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            f"""
            INSERT INTO vendas_parametros (cod_empresa, {coluna})
            VALUES (%s, %s)
            ON CONFLICT (cod_empresa)
            DO UPDATE SET
                {coluna} = EXCLUDED.{coluna},
                atualizado_em = now()
            """,
            (cod_empresa, sistema)
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def gravar_sistema_origem_painel(cod_empresa, sistema):
    _gravar_origem_vendas(cod_empresa, "sistema_origem_painel", sistema)


def gravar_sistema_origem_diarias(cod_empresa, sistema):
    _gravar_origem_vendas(cod_empresa, "sistema_origem_diarias", sistema)


# =========================
# LEITURA DO CSV DO O CLOSET
# =========================
# Colunas do relatório "vendas item por pagamento": uma linha por item vendido.
COLUNAS_CSV_OCLOSET = {
    "data": "Data",
    "preco_unitario": "Preço unitário",
    "total_item": "Total do item",
    "custo_total": "Custo total",
}


def ler_csv_ocloset(conteudo_bytes, data_inicial=None, data_final=None):
    """Resume o CSV de itens do O Closet no que o painel precisa.

    - quantidade: uma peça por linha do arquivo
    - valor: soma de "Total do item"
    - custo: soma de "Custo total"; quando vier vazio ou zero, metade do
      valor do item (o equivalente a metade do preço unitário por peça)
    - dias: quantidade de datas distintas presentes no arquivo

    O relatório do O Closet costuma vir com um intervalo maior que o mês que
    se quer importar; linhas com data fora de ``data_inicial``/``data_final``
    são descartadas (quando informadas) e contadas em ``linhas_fora_periodo``.
    """
    texto = None
    for codificacao in ("utf-8-sig", "latin-1"):
        try:
            texto = conteudo_bytes.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue

    if texto is None:
        raise ValueError("Não foi possível ler o arquivo (codificação desconhecida).")

    leitor = csv.DictReader(io.StringIO(texto), delimiter=";")

    cabecalho = [str(c or "").strip() for c in (leitor.fieldnames or [])]
    faltando = [
        nome for nome in COLUNAS_CSV_OCLOSET.values()
        if nome not in cabecalho
    ]
    if faltando:
        raise ValueError(
            "O arquivo não tem as colunas esperadas do O Closet: "
            + ", ".join(faltando)
        )

    quantidade = 0
    valor = Decimal("0")
    custo = Decimal("0")
    custo_estimado_linhas = 0
    linhas_fora_periodo = 0
    datas = set()

    for linha in leitor:
        if linha is None:
            continue

        data_txt = str(linha.get(COLUNAS_CSV_OCLOSET["data"]) or "").strip()
        total_txt = str(linha.get(COLUNAS_CSV_OCLOSET["total_item"]) or "").strip()

        # Linha totalmente em branco no fim do arquivo
        if not data_txt and not total_txt:
            continue

        try:
            data_linha = datetime.strptime(data_txt, "%d/%m/%Y").date()
        except ValueError:
            raise ValueError(f"Data inválida no arquivo: {data_txt!r}")

        if (data_inicial and data_linha < data_inicial) or \
           (data_final and data_linha > data_final):
            linhas_fora_periodo += 1
            continue

        datas.add(data_linha)

        valor_item = para_decimal(total_txt)
        custo_item = para_decimal(linha.get(COLUNAS_CSV_OCLOSET["custo_total"]))

        if custo_item <= 0:
            custo_item = (valor_item / Decimal("2")).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            custo_estimado_linhas += 1

        quantidade += 1
        valor += valor_item
        custo += custo_item

    if quantidade == 0:
        if linhas_fora_periodo:
            raise ValueError(
                "Nenhuma venda do arquivo está dentro do período informado "
                f"({linhas_fora_periodo} linhas ficaram fora)."
            )
        raise ValueError("O arquivo não tem linhas de venda.")

    return {
        "quantidade": quantidade,
        "valor": valor,
        "custo": custo,
        "margem_bruta": valor - custo,
        "dias": len(datas),
        "data_inicial": min(datas),
        "data_final": max(datas),
        "custo_estimado_linhas": custo_estimado_linhas,
        "linhas_fora_periodo": linhas_fora_periodo,
    }


# Colunas usadas na leitura diária (item a item, agrupado por dia x produto).
COLUNA_CSV_OCLOSET_PRODUTO = "Produto"

DIAS_SEMANA_PT = {
    0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta",
    4: "Sexta", 5: "Sábado", 6: "Domingo",
}


def ler_csv_ocloset_diario(conteudo_bytes):
    """Lê o CSV de itens do O Closet agrupando por (data, produto).

    Mesma regra de custo da importação do painel: item sem preço de compra
    entra com metade do valor de venda. Uma linha do arquivo = uma peça.

    Devolve (linhas, resumo), onde cada linha é um dicionário com data,
    descricao, quantidade, valor, custo e margem_bruta.
    """
    texto = None
    for codificacao in ("utf-8-sig", "latin-1"):
        try:
            texto = conteudo_bytes.decode(codificacao)
            break
        except UnicodeDecodeError:
            continue

    if texto is None:
        raise ValueError("Não foi possível ler o arquivo (codificação desconhecida).")

    leitor = csv.DictReader(io.StringIO(texto), delimiter=";")

    cabecalho = [str(c or "").strip() for c in (leitor.fieldnames or [])]
    esperadas = [
        COLUNAS_CSV_OCLOSET["data"],
        COLUNAS_CSV_OCLOSET["total_item"],
        COLUNAS_CSV_OCLOSET["custo_total"],
        COLUNA_CSV_OCLOSET_PRODUTO,
    ]
    faltando = [nome for nome in esperadas if nome not in cabecalho]
    if faltando:
        raise ValueError(
            "O arquivo não tem as colunas esperadas do O Closet: "
            + ", ".join(faltando)
        )

    agrupado = {}
    custo_estimado_linhas = 0
    total_linhas = 0

    for linha in leitor:
        if linha is None:
            continue

        data_txt = str(linha.get(COLUNAS_CSV_OCLOSET["data"]) or "").strip()
        total_txt = str(linha.get(COLUNAS_CSV_OCLOSET["total_item"]) or "").strip()

        # Linha totalmente em branco no fim do arquivo
        if not data_txt and not total_txt:
            continue

        try:
            data_linha = datetime.strptime(data_txt, "%d/%m/%Y").date()
        except ValueError:
            raise ValueError(f"Data inválida no arquivo: {data_txt!r}")

        descricao = str(linha.get(COLUNA_CSV_OCLOSET_PRODUTO) or "").strip()
        if not descricao:
            descricao = "(sem descrição)"
        descricao = descricao[:100]

        valor_item = para_decimal(total_txt)
        custo_item = para_decimal(linha.get(COLUNAS_CSV_OCLOSET["custo_total"]))

        if custo_item <= 0:
            custo_item = (valor_item / Decimal("2")).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            custo_estimado_linhas += 1

        chave = (data_linha, descricao)
        acumulado = agrupado.get(chave)

        if acumulado is None:
            agrupado[chave] = {
                "data": data_linha,
                "descricao": descricao,
                "quantidade": 1,
                "valor": valor_item,
                "custo": custo_item,
            }
        else:
            acumulado["quantidade"] += 1
            acumulado["valor"] += valor_item
            acumulado["custo"] += custo_item

        total_linhas += 1

    if not agrupado:
        raise ValueError("O arquivo não tem linhas de venda.")

    linhas = []
    for item in agrupado.values():
        item["margem_bruta"] = item["valor"] - item["custo"]
        linhas.append(item)

    linhas.sort(key=lambda x: (x["data"], x["descricao"]))

    datas = {item["data"] for item in linhas}

    resumo = {
        "linhas_arquivo": total_linhas,
        "custo_estimado_linhas": custo_estimado_linhas,
        "data_inicial": min(datas),
        "data_final": max(datas),
        "dias": len(datas),
        "quantidade": sum(item["quantidade"] for item in linhas),
        "valor": sum((item["valor"] for item in linhas), Decimal("0")),
        "custo": sum((item["custo"] for item in linhas), Decimal("0")),
    }
    resumo["margem_bruta"] = resumo["valor"] - resumo["custo"]

    return linhas, resumo


def mapear_codigos_produto_ocloset(cur, cod_empresa, descricoes):
    """Código inteiro estável por descrição de produto do O Closet.

    vendas_diarias.codigo_produto é inteiro e o CSV não traz código; a tabela
    vendas_produtos_ocloset guarda a correspondência para o mesmo produto cair
    sempre no mesmo código nas próximas importações.
    """
    cur.execute("""
        SELECT descricao, codigo_produto
        FROM vendas_produtos_ocloset
        WHERE cod_empresa = %s
    """, (cod_empresa,))
    mapa = {str(d): int(c) for d, c in cur.fetchall()}

    novas = [d for d in sorted(set(descricoes)) if d not in mapa]

    if novas:
        proximo = (max(mapa.values()) + 1) if mapa else 1
        registros = []

        for descricao in novas:
            mapa[descricao] = proximo
            registros.append((cod_empresa, descricao, proximo))
            proximo += 1

        execute_values(
            cur,
            """
            INSERT INTO vendas_produtos_ocloset (cod_empresa, descricao, codigo_produto)
            VALUES %s
            ON CONFLICT (cod_empresa, descricao) DO NOTHING
            """,
            registros
        )

    return mapa


def obter_nome_mes_abrev(mes):
    nomes = {
        1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr",
        5: "Mai", 6: "Jun", 7: "Jul", 8: "Ago",
        9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
    }
    return nomes.get(int(mes), str(mes))


def obter_parametros_padrao_importacao():
    hoje = date.today()

    if hoje.day == 1:
        referencia = hoje - timedelta(days=1)
    else:
        referencia = hoje

    ontem = hoje - timedelta(days=1)

    ano = referencia.year
    mes = referencia.month

    if hoje.day == 1:
        dia_base = referencia.day
    else:
        dia_base = ontem.day

    dias_mes = calendar.monthrange(ano, mes)[1]

    return {
        "ano": ano,
        "mes": mes,
        "dia_base": dia_base,
        "dias_mes": dias_mes
    }


def obter_periodo_padrao_consulta():
    hoje = datetime.now(ZoneInfo("America/Fortaleza")).date()
    data_fim = hoje
    data_ini = hoje - timedelta(days=30)
    return data_ini, data_fim


def obter_filiais_ativas(cur, cod_empresa):
    cur.execute("""
        SELECT cod_filial, nome_filial
        FROM filiais
        WHERE cod_empresa = %s
          AND ativo = TRUE
        ORDER BY cod_filial
    """, (cod_empresa,))
    return cur.fetchall()


def montar_grade_sintetica(filiais, registros, campo_valor):
    meses_ordenados = sorted(
        {(int(r["ano"]), int(r["mes"])) for r in registros},
        key=lambda x: (x[0], x[1])
    )

    meses_ordenados = meses_ordenados[-24:]
    meses_validos = set(meses_ordenados)

    mapa = {}
    for r in registros:
        ano = int(r["ano"])
        mes = int(r["mes"])

        if (ano, mes) not in meses_validos:
            continue

        chave = (ano, mes, int(r["cod_filial"]))
        mapa[chave] = float(r[campo_valor] or 0)

    linhas = []
    totais_por_filial = defaultdict(float)
    serie_por_filial = defaultdict(list)

    for ano, mes in meses_ordenados:
        linha = {
            "periodo": f"{obter_nome_mes_abrev(mes)}/{str(ano)[-2:]}",
            "ano": ano,
            "mes": mes,
            "total": 0.0,
            "valores": []
        }

        for filial in filiais:
            cod_filial = int(filial["cod_filial"])
            valor = mapa.get((ano, mes, cod_filial))
            linha["valores"].append(valor)

            if valor not in (None, 0, 0.0, ""):
                linha["total"] += valor
                totais_por_filial[cod_filial] += valor
                serie_por_filial[cod_filial].append(valor)

        linhas.append(linha)

    total_geral = sum(l["total"] for l in linhas)

    linha_total = {
        "rotulo": "TOTAL",
        "valores": [totais_por_filial[int(f["cod_filial"])] for f in filiais],
        "total": total_geral
    }

    linha_med_12m = {
        "rotulo": "MED 12 M",
        "valores": [],
        "total": 0.0
    }

    for filial in filiais:
        cod_filial = int(filial["cod_filial"])
        serie = serie_por_filial[cod_filial][-12:]
        media = sum(serie) / len(serie) if serie else 0.0
        linha_med_12m["valores"].append(media)

    linha_med_12m["total"] = sum(linha_med_12m["valores"])

    linha_proj_12m = {
        "rotulo": "PROJ 12 M",
        "valores": [v * 12 for v in linha_med_12m["valores"]],
        "total": sum(v * 12 for v in linha_med_12m["valores"])
    }

    return {
        "linhas": linhas,
        "linha_total": linha_total,
        "linha_med_12m": linha_med_12m,
        "linha_proj_12m": linha_proj_12m
    }


def montar_grade_mun(filiais, grade_mb, grade_unidades):
    mapa_mb = {(linha["ano"], linha["mes"]): linha for linha in grade_mb["linhas"]}
    mapa_un = {(linha["ano"], linha["mes"]): linha for linha in grade_unidades["linhas"]}

    chaves = sorted(set(mapa_mb.keys()) | set(mapa_un.keys()))
    linhas = []

    totais_mun_filial = []
    med_mun_filial = []
    proj_mun_filial = []

    for i, _filial in enumerate(filiais):
        total_mb = grade_mb["linha_total"]["valores"][i]
        total_un = grade_unidades["linha_total"]["valores"][i]
        totais_mun_filial.append((total_mb / total_un) if total_un else 0.0)

        med_mb = grade_mb["linha_med_12m"]["valores"][i]
        med_un = grade_unidades["linha_med_12m"]["valores"][i]
        med_mun_filial.append((med_mb / med_un) if med_un else 0.0)

        proj_mb = grade_mb["linha_proj_12m"]["valores"][i]
        proj_un = grade_unidades["linha_proj_12m"]["valores"][i]
        proj_mun_filial.append((proj_mb / proj_un) if proj_un else 0.0)

    for ano, mes in chaves:
        linha_mb = mapa_mb.get((ano, mes))
        linha_un = mapa_un.get((ano, mes))

        valores = []
        total_mb_linha = linha_mb["total"] if linha_mb else None
        total_un_linha = linha_un["total"] if linha_un else None

        for i in range(len(filiais)):
            mb = linha_mb["valores"][i] if linha_mb else None
            un = linha_un["valores"][i] if linha_un else None

            if mb in (None, 0, 0.0, "") or un in (None, 0, 0.0, ""):
                valores.append(None)
            else:
                valores.append(mb / un)

        linhas.append({
            "periodo": f"{obter_nome_mes_abrev(mes)}/{str(ano)[-2:]}",
            "ano": ano,
            "mes": mes,
            "valores": valores,
            "total": (total_mb_linha / total_un_linha)
            if total_mb_linha not in (None, 0, 0.0, "") and total_un_linha not in (None, 0, 0.0, "")
            else None
        })

    return {
        "linhas": linhas,
        "linha_total": {
            "rotulo": "TOTAL",
            "valores": totais_mun_filial,
            "total": (
                grade_mb["linha_total"]["total"] / grade_unidades["linha_total"]["total"]
                if grade_unidades["linha_total"]["total"] else 0.0
            )
        },
        "linha_med_12m": {
            "rotulo": "MED 12 M",
            "valores": med_mun_filial,
            "total": (
                grade_mb["linha_med_12m"]["total"] / grade_unidades["linha_med_12m"]["total"]
                if grade_unidades["linha_med_12m"]["total"] else 0.0
            )
        },
        "linha_proj_12m": {
            "rotulo": "PROJ 12 M",
            "valores": proj_mun_filial,
            "total": (
                grade_mb["linha_proj_12m"]["total"] / grade_unidades["linha_proj_12m"]["total"]
                if grade_unidades["linha_proj_12m"]["total"] else 0.0
            )
        }
    }


def montar_resumo_projecao(grade_unidades, grade_valores, grade_mb, dias_mes=None):
    if not grade_unidades["linhas"]:
        return None

    ultima_qtd = grade_unidades["linhas"][-1]
    ultima_val = grade_valores["linhas"][-1] if grade_valores["linhas"] else None
    ultima_mb = grade_mb["linhas"][-1] if grade_mb["linhas"] else None

    quantidade = float(ultima_qtd.get("total") or 0)
    valor = float(ultima_val.get("total") or 0) if ultima_val else 0.0
    mb = float(ultima_mb.get("total") or 0) if ultima_mb else 0.0

    mun = (mb / quantidade) if quantidade else 0.0
    qtd_dia = (quantidade / dias_mes) if dias_mes else 0.0
    mb_dia = (mb / dias_mes) if dias_mes else 0.0

    return {
        "titulo": f"PROJEÇÃO {ultima_qtd['periodo'].upper()}",
        "periodo": ultima_qtd["periodo"],
        "quantidade": quantidade,
        "valor": valor,
        "mb": mb,
        "mun": mun,
        "qtd_dia": qtd_dia,
        "mb_dia": mb_dia,
        "dias_mes": dias_mes
    }


def montar_grade_variacoes_projecoes(registros):
    if not registros:
        return []

    linhas = []

    for r in registros:
        quantidade = float(r["quantidade_proj"]) if r.get("quantidade_proj") not in (None, "") else 0.0
        valor = float(r["valor_proj"]) if r.get("valor_proj") not in (None, "") else 0.0
        mb = float(r["mb_proj"]) if r.get("mb_proj") not in (None, "") else 0.0
        mun = (mb / quantidade) if quantidade else 0.0

        data_importacao = r.get("data_importacao")
        data_fmt = formatar_data_brasil(data_importacao)

        linhas.append({
            "data": data_fmt,
            "dia_base": int(r["dia_base"]) if r.get("dia_base") not in (None, "") else 0,
            "dias_mes": int(r["dias_mes"]) if r.get("dias_mes") not in (None, "") else 0,
            "quantidade": quantidade,
            "valor": valor,
            "mb": mb,
            "mun": mun
        })

    linhas = aplicar_heatmap_variacoes(
        linhas,
        ["quantidade", "valor", "mb", "mun"]
    )

    return linhas

def criar_job_importacao(cur, cod_empresa, tipo_importacao):
    job_id = str(uuid.uuid4())
    cur.execute("""
        INSERT INTO importacoes_progresso (
            job_id,
            cod_empresa,
            tipo_importacao,
            status,
            etapa,
            mensagem,
            percentual,
            total_linhas,
            linhas_processadas
        )
        VALUES (%s, %s, %s, 'processando', '', '', 0, 0, 0)
    """, (job_id, cod_empresa, tipo_importacao))
    return job_id


def atualizar_job_importacao(
    cur,
    job_id,
    status=None,
    etapa=None,
    mensagem=None,
    percentual=None,
    total_linhas=None,
    linhas_processadas=None
):
    sets = ["atualizado_em = NOW()"]
    params = []

    if status is not None:
        sets.append("status = %s")
        params.append(status)

    if etapa is not None:
        sets.append("etapa = %s")
        params.append(etapa)

    if mensagem is not None:
        sets.append("mensagem = %s")
        params.append(mensagem)

    if percentual is not None:
        sets.append("percentual = %s")
        params.append(int(percentual))

    if total_linhas is not None:
        sets.append("total_linhas = %s")
        params.append(int(total_linhas))

    if linhas_processadas is not None:
        sets.append("linhas_processadas = %s")
        params.append(int(linhas_processadas))

    params.append(job_id)

    cur.execute(f"""
        UPDATE importacoes_progresso
           SET {", ".join(sets)}
         WHERE job_id = %s
    """, params)

# =========================
# PAINEL
# =========================
@vendas_bp.route("/painel")
@permissao_obrigatoria("VENDAS", "CONSULTAR_PAINEL", redirecionar_para="sistema.menu_vendas")
def vendas_painel():
    if "id_usuario" not in session:
        return redirect(url_for("auth.index"))

    if "cod_empresa" not in session:
        return redirect(url_for("auth.escolher_empresa"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT cod_filial, nome_filial
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY cod_filial
        """, (cod_empresa,))
        filiais = cur.fetchall()

        cur.execute("""
            SELECT cod_empresa, cod_filial, ano, mes, quantidade_vendida
            FROM vendas_unidades_sintetico
            WHERE cod_empresa = %s
            ORDER BY ano, mes, cod_filial
        """, (cod_empresa,))
        registros_unidades = cur.fetchall()

        cur.execute("""
            SELECT cod_empresa, cod_filial, ano, mes, valor_vendido
            FROM vendas_valores_sintetico
            WHERE cod_empresa = %s
            ORDER BY ano, mes, cod_filial
        """, (cod_empresa,))
        registros_valores = cur.fetchall()

        cur.execute("""
            SELECT cod_empresa, cod_filial, ano, mes, margem_bruta
            FROM vendas_mb_sintetico
            WHERE cod_empresa = %s
            ORDER BY ano, mes, cod_filial
        """, (cod_empresa,))
        registros_mb = cur.fetchall()

        cur.execute("""
            SELECT ano, mes, dia_base, dias_mes, data_importacao
            FROM vendas_painel_importacoes
            WHERE cod_empresa = %s
            ORDER BY data_importacao DESC
            LIMIT 1
        """, (cod_empresa,))
        meta_importacao = cur.fetchone()

        cur.execute("""
            SELECT
                ano,
                mes,
                dia_base,
                dias_mes,
                quantidade_proj,
                valor_proj,
                mb_proj,
                data_importacao
            FROM vendas_painel_importacoes
            WHERE cod_empresa = %s
            ORDER BY data_importacao DESC
            LIMIT 30
        """, (cod_empresa,))
        registros_variacoes = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    grade_unidades = montar_grade_sintetica(filiais, registros_unidades, "quantidade_vendida")
    grade_valores = montar_grade_sintetica(filiais, registros_valores, "valor_vendido")
    grade_mb = montar_grade_sintetica(filiais, registros_mb, "margem_bruta")
    grade_mun = montar_grade_mun(filiais, grade_mb, grade_unidades)

    grade_unidades = aplicar_heatmap_na_grade(grade_unidades)
    grade_valores = aplicar_heatmap_na_grade(grade_valores)
    grade_mb = aplicar_heatmap_na_grade(grade_mb)
    grade_mun = aplicar_heatmap_na_grade(grade_mun)

    grade_unidades = aplicar_heatmap_na_coluna_total(grade_unidades)
    grade_valores = aplicar_heatmap_na_coluna_total(grade_valores)
    grade_mb = aplicar_heatmap_na_coluna_total(grade_mb)
    grade_mun = aplicar_heatmap_na_coluna_total(grade_mun)

    dias_mes_resumo = meta_importacao["dias_mes"] if meta_importacao else None

    resumo_projecao = montar_resumo_projecao(
        grade_unidades,
        grade_valores,
        grade_mb,
        dias_mes=dias_mes_resumo
    )

    variacoes_projecoes = montar_grade_variacoes_projecoes(
        list(reversed(registros_variacoes))
    )

    return render_template(
        "vendas_painel.html",
        nome_empresa=nome_empresa,
        filiais=filiais,
        grade_unidades=grade_unidades,
        grade_valores=grade_valores,
        grade_mb=grade_mb,
        grade_mun=grade_mun,
        resumo_projecao=resumo_projecao,
        variacoes_projecoes=variacoes_projecoes,
        formatar_numero_br=formatar_numero_br,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar"
    )


# =========================
# PARÂMETROS DE VENDAS
# =========================
@vendas_bp.route("/parametros", methods=["GET", "POST"])
@permissao_obrigatoria("VENDAS", "PARAMETROS", redirecionar_para="sistema.menu_vendas")
def vendas_parametros():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()

    if request.method == "POST":
        sistema = str(request.form.get("sistema_origem_painel") or "").strip().upper()
        sistema_diarias = str(request.form.get("sistema_origem_diarias") or "").strip().upper()

        if sistema not in SISTEMAS_ORIGEM_PAINEL or sistema_diarias not in SISTEMAS_ORIGEM_DIARIAS:
            flash("Selecione um sistema de origem válido.", "error")
        else:
            try:
                gravar_sistema_origem_painel(cod_empresa, sistema)
                gravar_sistema_origem_diarias(cod_empresa, sistema_diarias)
                g.pop("_origem_painel", None)
                g.pop("_origem_diarias", None)
                flash("Parâmetros salvos.", "success")
            except Exception as e:
                flash(f"Erro ao salvar: {e}", "error")

        return redirect(url_for("vendas.vendas_parametros"))

    return render_template(
        "vendas_parametros.html",
        nome_empresa=session.get("nome_empresa", ""),
        sistemas_origem=SISTEMAS_ORIGEM_PAINEL,
        sistema_origem_atual=obter_sistema_origem_painel(cod_empresa),
        sistemas_origem_diarias=SISTEMAS_ORIGEM_DIARIAS,
        sistema_origem_diarias_atual=obter_sistema_origem_diarias(cod_empresa),
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar"
    )


# =========================
# IMPORTAÇÃO DO PAINEL — O CLOSET
# =========================
def importar_painel_ocloset(cod_empresa, nome_empresa, padrao):
    """Importa o CSV de itens do O Closet para o painel sintético.

    O painel guarda valores projetados para o mês inteiro; aqui a média diária
    (dividida pelos dias distintos do arquivo) é multiplicada pelos dias do mês.
    """
    hoje = date.today()

    def pagina(**kw):
        base = {
            "nome_empresa": nome_empresa,
            "ano_sugerido": padrao["ano"],
            "mes_sugerido": padrao["mes"],
            "dias_mes_sugerido": padrao["dias_mes"],
            # Corte padrão: do primeiro dia do mês atual até ontem — o dia de
            # hoje ainda está em andamento e entraria pela metade. O relatório
            # do O Closet costuma vir com um intervalo maior.
            "data_inicial_sugerida": hoje.replace(day=1).isoformat(),
            "data_final_sugerida": (hoje - timedelta(days=1)).isoformat(),
            "url_voltar": url_for("sistema.menu_vendas"),
            "texto_voltar": "← Voltar",
        }
        base.update(kw)
        return render_template("vendas_importar_painel_ocloset.html", **base)

    if request.method == "GET":
        return pagina()

    arquivo = request.files.get("arquivo")
    ano_txt = (request.form.get("ano") or "").strip()
    mes_txt = (request.form.get("mes") or "").strip()
    dias_mes_txt = (request.form.get("dias_mes") or "").strip()
    data_inicial_txt = (request.form.get("data_inicial") or "").strip()
    data_final_txt = (request.form.get("data_final") or "").strip()

    # Mantém o que o usuário digitou quando a página é redesenhada
    eco = {
        "data_inicial_sugerida": data_inicial_txt,
        "data_final_sugerida": data_final_txt,
    }

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo.", "error")
        return pagina(**eco)

    if not (ano_txt.isdigit() and mes_txt.isdigit() and dias_mes_txt.isdigit()):
        flash("Informe ano, mês e dias do mês válidos.", "error")
        return pagina(**eco)

    try:
        data_inicial = date.fromisoformat(data_inicial_txt)
        data_final = date.fromisoformat(data_final_txt)
    except ValueError:
        flash("Informe a data inicial e a data final do corte.", "error")
        return pagina(**eco)

    if data_inicial > data_final:
        flash("A data inicial não pode ser maior que a data final.", "error")
        return pagina(**eco)

    eco.update({
        "ano_sugerido": int(ano_txt),
        "mes_sugerido": int(mes_txt),
        "dias_mes_sugerido": int(dias_mes_txt),
    })

    ano = int(ano_txt)
    mes = int(mes_txt)
    dias_mes = int(dias_mes_txt)

    if mes < 1 or mes > 12:
        flash("O mês deve estar entre 1 e 12.", "error")
        return pagina(**eco)

    if dias_mes < 1 or dias_mes > 31:
        flash("Dias do mês inválido.", "error")
        return pagina(**eco)

    try:
        resumo = ler_csv_ocloset(arquivo.read(), data_inicial, data_final)
    except Exception as e:
        flash(f"Erro ao ler o arquivo: {e}", "error")
        return pagina(**eco)

    # Base da média: dias CORRIDOS do corte, da data inicial até o último dia
    # com venda no arquivo. Dia sem venda nenhuma também é dia decorrido do mês
    # e precisa entrar no divisor — usar só os dias com venda inflava a média.
    dias = (resumo["data_final"] - data_inicial).days + 1
    dias_com_venda = resumo["dias"]
    quantidade_dia = Decimal(resumo["quantidade"]) / Decimal(dias)
    valor_dia = resumo["valor"] / Decimal(dias)
    mb_dia = resumo["margem_bruta"] / Decimal(dias)

    def projetar(valor_diario):
        return (valor_diario * Decimal(dias_mes)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )

    quantidade_proj = projetar(quantidade_dia)
    valor_proj = projetar(valor_dia)
    mb_proj = projetar(mb_dia)

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT cod_filial
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY cod_filial
        """, (cod_empresa,))
        filiais = cur.fetchall()

        if not filiais:
            flash("Nenhuma filial ativa cadastrada para a empresa.", "error")
            return pagina(**eco)

        if len(filiais) > 1:
            flash(
                "A empresa tem mais de uma filial ativa; o arquivo do O Closet não "
                "separa por filial e tudo foi lançado na primeira "
                f"(filial {filiais[0]['cod_filial']}).",
                "warning"
            )

        cod_filial = filiais[0]["cod_filial"]

        cur2 = conn.cursor()

        limpar_importacao_mes(cur2, cod_empresa, ano, mes)

        upsert(cur2, "vendas_unidades_sintetico", "quantidade_vendida",
               cod_empresa, cod_filial, ano, mes, float(quantidade_proj))
        upsert(cur2, "vendas_valores_sintetico", "valor_vendido",
               cod_empresa, cod_filial, ano, mes, float(valor_proj))
        upsert(cur2, "vendas_mb_sintetico", "margem_bruta",
               cod_empresa, cod_filial, ano, mes, float(mb_proj))

        inserir_importacao_painel(
            cur2, cod_empresa, ano, mes, dias, dias_mes,
            quantidade_proj, valor_proj, mb_proj
        )

        conn.commit()
        cur2.close()

        flash(
            "Conferência da importação — "
            f"Período: {resumo['data_inicial'].strftime('%d/%m/%Y')} a "
            f"{resumo['data_final'].strftime('%d/%m/%Y')} "
            f"({dias} dias corridos, {dias_com_venda} com venda) | "
            f"Qtd lida: {resumo['quantidade']} | "
            f"Vlr lido: {formatar_numero_br(float(resumo['valor']))} | "
            f"Custo: {formatar_numero_br(float(resumo['custo']))} | "
            f"MB lida: {formatar_numero_br(float(resumo['margem_bruta']))} | "
            f"Qtd/dia: {formatar_numero_br(float(quantidade_dia))} | "
            f"MB/dia: {formatar_numero_br(float(mb_dia))}",
            "warning"
        )

        if resumo["linhas_fora_periodo"]:
            flash(
                f"{resumo['linhas_fora_periodo']} linhas do arquivo estavam fora do "
                f"corte {data_inicial.strftime('%d/%m/%Y')} a "
                f"{data_final.strftime('%d/%m/%Y')} e foram desconsideradas.",
                "warning"
            )

        if resumo["custo_estimado_linhas"]:
            flash(
                f"{resumo['custo_estimado_linhas']} itens estavam sem preço de compra — "
                "o custo desses itens foi estimado em metade do valor de venda.",
                "warning"
            )

        flash(
            "Importação concluída. Projeção para o mês: "
            f"Qtd {formatar_numero_br(float(quantidade_proj))} | "
            f"Vlr {formatar_numero_br(float(valor_proj))} | "
            f"MB {formatar_numero_br(float(mb_proj))} "
            f"(média de {dias} dias corridos x {dias_mes} dias do mês).",
            "success"
        )

    except Exception as e:
        conn.rollback()
        flash(f"Erro: {e}", "error")

    finally:
        cur.close()
        conn.close()

    return pagina(**eco)


# =========================
# IMPORTAÇÃO DO PAINEL
# =========================
@vendas_bp.route("/painel/importar", methods=["GET", "POST"])
@permissao_obrigatoria("VENDAS", "IMPORTAR_PAINEL", redirecionar_para="sistema.menu_vendas")
def vendas_importar_painel():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")
    padrao = obter_parametros_padrao_importacao()

    if obter_sistema_origem_painel(cod_empresa) == "OCLOSET":
        return importar_painel_ocloset(cod_empresa, nome_empresa, padrao)

    if request.method == "GET":
        return render_template(
            "vendas_importar_painel.html",
            nome_empresa=nome_empresa,
            ano_sugerido=padrao["ano"],
            mes_sugerido=padrao["mes"],
            dia_base_sugerido=padrao["dia_base"],
            dias_mes_sugerido=padrao["dias_mes"],
            url_voltar=url_for("sistema.menu_vendas"),
            texto_voltar="← Voltar"
        )

    arquivo = request.files.get("arquivo")
    ano_txt = (request.form.get("ano") or "").strip()
    mes_txt = (request.form.get("mes") or "").strip()
    dia_base_txt = (request.form.get("dia_base") or "").strip()
    dias_mes_txt = (request.form.get("dias_mes") or "").strip()


    
    arquivo = request.files.get("arquivo")
    ano_txt = (request.form.get("ano") or "").strip()
    mes_txt = (request.form.get("mes") or "").strip()
    dia_base_txt = (request.form.get("dia_base") or "").strip()
    dias_mes_txt = (request.form.get("dias_mes") or "").strip()

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo.", "error")
        return render_template(
            "vendas_importar_painel.html",
            nome_empresa=nome_empresa,
            ano_sugerido=padrao["ano"],
            mes_sugerido=padrao["mes"],
            dia_base_sugerido=padrao["dia_base"],
            dias_mes_sugerido=padrao["dias_mes"]
        )

    if not (ano_txt.isdigit() and mes_txt.isdigit() and dia_base_txt.isdigit() and dias_mes_txt.isdigit()):
        flash("Informe ano, mês, dia base e dias do mês válidos.", "error")
        return render_template(
            "vendas_importar_painel.html",
            nome_empresa=nome_empresa,
            ano_sugerido=padrao["ano"],
            mes_sugerido=padrao["mes"],
            dia_base_sugerido=padrao["dia_base"],
            dias_mes_sugerido=padrao["dias_mes"]
        )

    ano = int(ano_txt)
    mes = int(mes_txt)
    dia_base = int(dia_base_txt)
    dias_mes = int(dias_mes_txt)

    if mes < 1 or mes > 12:
        flash("O mês deve estar entre 1 e 12.", "error")
        return render_template(
            "vendas_importar_painel.html",
            nome_empresa=nome_empresa,
            ano_sugerido=padrao["ano"],
            mes_sugerido=padrao["mes"],
            dia_base_sugerido=padrao["dia_base"],
            dias_mes_sugerido=padrao["dias_mes"]
        )

    ultimo_dia_mes = calendar.monthrange(ano, mes)[1]

    if dia_base < 1 or dia_base > ultimo_dia_mes:
        flash(f"O dia base deve estar entre 1 e {ultimo_dia_mes}.", "error")
        return render_template(
            "vendas_importar_painel.html",
            nome_empresa=nome_empresa,
            ano_sugerido=ano,
            mes_sugerido=mes,
            dia_base_sugerido=dia_base,
            dias_mes_sugerido=dias_mes
        )

    if dias_mes < 1 or dias_mes > 31:
        flash("Dias do mês inválido.", "error")
        return render_template(
            "vendas_importar_painel.html",
            nome_empresa=nome_empresa,
            ano_sugerido=ano,
            mes_sugerido=mes,
            dia_base_sugerido=dia_base,
            dias_mes_sugerido=dias_mes
        )

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    tmp = None

    try:
        cur.execute("""
            SELECT cod_filial, nome_filial_importacao
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
        """, (cod_empresa,))
        filiais = cur.fetchall()

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        arquivo.save(tmp.name)
        tmp.close()

        wb = load_workbook(tmp.name, data_only=True)
        ws = wb.worksheets[0]

        # --- validação prévia de filiais ---
        # Coleta nomes de filiais presentes no Excel (células col A imediatamente antes de "TOTAL FILIAL:")
        nomes_painel_excel = set()
        ultima_col_a = None
        for rl in range(1, 10000):
            val_cel = ws.cell(row=rl, column=1).value
            txt_cel = str(val_cel).strip().upper() if val_cel is not None else ""
            if txt_cel == "TOTAL FILIAL:":
                if ultima_col_a:
                    nomes_painel_excel.add(ultima_col_a)
            if txt_cel:
                ultima_col_a = txt_cel

        filiais_nao_encontradas = []
        for f in filiais:
            nome_busca = f["nome_filial_importacao"]
            if not nome_busca:
                continue
            if not localizar_total_filial(ws, nome_busca):
                filiais_nao_encontradas.append(nome_busca)

        # Filiais no Excel que não batem com nenhuma do cadastro
        nomes_db_upper = [
            str(f["nome_filial_importacao"]).strip().upper()
            for f in filiais
            if f["nome_filial_importacao"]
        ]
        excel_sem_cadastro = [
            n for n in sorted(nomes_painel_excel)
            if not any(db in n or n in db for db in nomes_db_upper)
            and n not in ("TOTAL FILIAL:", "")
        ]

        if filiais_nao_encontradas:
            flash(
                "ATENÇÃO — As seguintes filiais do cadastro NÃO foram encontradas no arquivo Excel e NÃO serão importadas: "
                + ", ".join(filiais_nao_encontradas),
                "error"
            )
        if excel_sem_cadastro:
            flash(
                "ATENÇÃO — As seguintes entradas do Excel NÃO correspondem a nenhuma filial do cadastro: "
                + ", ".join(excel_sem_cadastro),
                "warning"
            )
        # --- fim validação ---

        cur2 = conn.cursor()

        limpar_importacao_mes(cur2, cod_empresa, ano, mes)

        importadas = 0

        total_unidades_lidas = Decimal("0")
        total_valor_lido = Decimal("0")
        total_mb_lido = Decimal("0")

        total_unidades_proj = Decimal("0")
        total_valor_proj = Decimal("0")
        total_mb_proj = Decimal("0")

        for f in filiais:
            cod_filial = f["cod_filial"]
            nome_busca = f["nome_filial_importacao"]

            if not nome_busca:
                continue

            linha = localizar_total_filial(ws, nome_busca)
            if not linha:
                continue

            unidades_lidas = para_decimal(ws.cell(row=linha, column=7).value)
            valor_lido = para_decimal(ws.cell(row=linha, column=9).value)
            mb_lido = para_decimal(ws.cell(row=linha, column=11).value)

            fator_proj = (
                Decimal(str(dias_mes)) / Decimal(str(dia_base))
                if dia_base > 0 else Decimal("1")
            )

            unidades_proj = (unidades_lidas * fator_proj).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            valor_proj = (valor_lido * fator_proj).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            mb_proj = (mb_lido * fator_proj).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            total_unidades_lidas += unidades_lidas
            total_valor_lido += valor_lido
            total_mb_lido += mb_lido

            total_unidades_proj += unidades_proj
            total_valor_proj += valor_proj
            total_mb_proj += mb_proj

            upsert(
                cur2,
                "vendas_unidades_sintetico",
                "quantidade_vendida",
                cod_empresa,
                cod_filial,
                ano,
                mes,
                float(unidades_proj)
            )

            upsert(
                cur2,
                "vendas_valores_sintetico",
                "valor_vendido",
                cod_empresa,
                cod_filial,
                ano,
                mes,
                float(valor_proj)
            )

            upsert(
                cur2,
                "vendas_mb_sintetico",
                "margem_bruta",
                cod_empresa,
                cod_filial,
                ano,
                mes,
                float(mb_proj)
            )

            importadas += 1

        inserir_importacao_painel(
            cur2,
            cod_empresa,
            ano,
            mes,
            dia_base,
            dias_mes,
            total_unidades_proj,
            total_valor_proj,
            total_mb_proj
        )

        flash(
            "Conferência da importação — "
            f"Qtd lida: {formatar_numero_br(float(total_unidades_lidas))} | "
            f"Vlr lido: {formatar_numero_br(float(total_valor_lido))} | "
            f"MB lida: {formatar_numero_br(float(total_mb_lido))} | "
            f"Qtd proj: {formatar_numero_br(float(total_unidades_proj))} | "
            f"Vlr proj: {formatar_numero_br(float(total_valor_proj))} | "
            f"MB proj: {formatar_numero_br(float(total_mb_proj))}",
            "warning"
        )

        conn.commit()
        cur2.close()

        flash(
            f"{importadas} filiais importadas com sucesso. "
            f"Projeção aplicada: dia base {dia_base} / dias do mês {dias_mes}.",
            "success"
        )

    except Exception as e:
        conn.rollback()
        flash(f"Erro: {e}", "error")

    finally:
        cur.close()
        conn.close()
        if tmp is not None and os.path.exists(tmp.name):
            os.unlink(tmp.name)

    return render_template(
        "vendas_importar_painel.html",
        nome_empresa=nome_empresa,
        ano_sugerido=ano,
        mes_sugerido=mes,
        dia_base_sugerido=dia_base,
        dias_mes_sugerido=dias_mes,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar para o Menu"
    )


# =========================
# VENDAS DIÁRIAS
# =========================
@vendas_bp.route("/diarias")
@permissao_obrigatoria("VENDAS", "CONSULTAR_VENDAS_DIARIAS", redirecionar_para="sistema.menu_vendas")
def vendas_diarias():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    data_ini_padrao, data_fim_padrao = obter_periodo_padrao_consulta()

    data_ini_txt = (request.args.get("data_ini") or "").strip()
    data_fim_txt = (request.args.get("data_fim") or "").strip()

    data_ini = para_data_excel(data_ini_txt) if data_ini_txt else data_ini_padrao
    data_fim = para_data_excel(data_fim_txt) if data_fim_txt else data_fim_padrao

    if not data_ini:
        data_ini = data_ini_padrao
    if not data_fim:
        data_fim = data_fim_padrao

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                data,
                dia_semana,
                SUM(quantidade) AS quantidade,
                SUM(valor) AS valor,
                SUM(margem_bruta) AS mb
            FROM vendas_diarias
            WHERE cod_empresa = %s
              AND data BETWEEN %s AND %s
            GROUP BY data, dia_semana
            ORDER BY data ASC
        """, (cod_empresa, data_ini, data_fim))
        linhas_totais = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    for linha in linhas_totais:
        quantidade = float(linha["quantidade"] or 0)
        valor = float(linha["valor"] or 0)
        mb = float(linha["mb"] or 0)

        linha["quantidade"] = quantidade
        linha["valor"] = valor
        linha["mb"] = mb
        linha["mun"] = (mb / quantidade) if quantidade else 0.0
        linha["data_fmt"] = formatar_data_brasil(linha["data"])

    linhas_totais = aplicar_heatmap_consulta(linhas_totais)

    return render_template(
        "vendas_diarias.html",
        nome_empresa=nome_empresa,
        data_ini=data_ini.strftime("%Y-%m-%d"),
        data_fim=data_fim.strftime("%Y-%m-%d"),
        linhas_totais=linhas_totais,
        formatar_numero_br=formatar_numero_br,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar"
    )
# =========================
# VENDAS DIÁRIAS IMPORTAR
# =========================
def importar_diarias_ocloset(cod_empresa, nome_empresa):
    """Importa o CSV de itens do O Closet para vendas_diarias.

    O arquivo é item a item; aqui vira uma linha por (dia, produto), no mesmo
    formato que a importação do WebPostos grava — só muda o tipo de produto.
    O período regravado é o próprio intervalo lido do arquivo.
    """
    def pagina(**kw):
        base = {
            "nome_empresa": nome_empresa,
            "url_voltar": url_for("sistema.menu_vendas"),
            "texto_voltar": "← Voltar",
        }
        base.update(kw)
        return render_template("vendas_importar_diarias_ocloset.html", **base)

    if request.method == "GET":
        return pagina()

    arquivo = request.files.get("arquivo")

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo.", "error")
        return pagina()

    try:
        linhas, resumo = ler_csv_ocloset_diario(arquivo.read())
    except Exception as e:
        flash(f"Erro ao ler o arquivo: {e}", "error")
        return pagina()

    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT cod_filial
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY cod_filial
        """, (cod_empresa,))
        filiais = cur.fetchall()

        if not filiais:
            flash("Nenhuma filial ativa cadastrada para a empresa.", "error")
            return pagina()

        if len(filiais) > 1:
            flash(
                "A empresa tem mais de uma filial ativa; o arquivo do O Closet não "
                "separa por filial e tudo foi lançado na primeira "
                f"(filial {filiais[0][0]}).",
                "warning"
            )

        cod_filial = int(filiais[0][0])

        mapa_codigos = mapear_codigos_produto_ocloset(
            cur, cod_empresa, [item["descricao"] for item in linhas]
        )

        dados = []
        for item in linhas:
            quantidade = Decimal(item["quantidade"])
            preco_venda = (item["valor"] / quantidade).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            custo_unitario = (item["custo"] / quantidade).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

            dados.append((
                cod_empresa,
                cod_filial,
                item["data"],
                DIAS_SEMANA_PT[item["data"].weekday()],
                mapa_codigos[item["descricao"]],
                item["descricao"],
                float(custo_unitario),
                float(preco_venda),
                item["quantidade"],
                float(item["valor"]),
                float(item["margem_bruta"]),
            ))

        limpar_importacao_diaria_periodo(
            cur, conn, cod_empresa, resumo["data_inicial"], resumo["data_final"]
        )

        execute_values(
            cur,
            """
            INSERT INTO vendas_diarias (
                cod_empresa,
                cod_filial,
                data,
                dia_semana,
                codigo_produto,
                descricao,
                custo,
                preco_venda,
                quantidade,
                valor,
                margem_bruta
            )
            VALUES %s
            """,
            dados,
            page_size=5000
        )

        conn.commit()

        flash(
            f"{len(dados)} registros importados (uma linha por dia e produto), "
            f"a partir de {resumo['linhas_arquivo']} itens do arquivo. "
            f"Período: {resumo['data_inicial'].strftime('%d/%m/%Y')} a "
            f"{resumo['data_final'].strftime('%d/%m/%Y')} "
            f"({resumo['dias']} dias com venda) | "
            f"Qtd: {resumo['quantidade']} | "
            f"Vlr: {formatar_numero_br(float(resumo['valor']))} | "
            f"Custo: {formatar_numero_br(float(resumo['custo']))} | "
            f"MB: {formatar_numero_br(float(resumo['margem_bruta']))}",
            "success"
        )

        if resumo["custo_estimado_linhas"]:
            flash(
                f"{resumo['custo_estimado_linhas']} itens estavam sem preço de compra — "
                "o custo desses itens foi estimado em metade do valor de venda.",
                "warning"
            )

    except Exception as e:
        conn.rollback()
        flash(f"Erro: {e}", "error")

    finally:
        cur.close()
        conn.close()

    return pagina()


@vendas_bp.route("/diarias/importar", methods=["GET", "POST"])
@permissao_obrigatoria("VENDAS", "IMPORTAR_VENDAS_DIARIAS", redirecionar_para="sistema.menu_vendas")
def vendas_importar_diarias():

    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    if obter_sistema_origem_diarias(cod_empresa) == "OCLOSET":
        return importar_diarias_ocloset(cod_empresa, nome_empresa)

    if request.method == "GET":
        return render_template(
            "vendas_importar_diarias.html",
            nome_empresa=nome_empresa,
            job_id=None,
            url_voltar=url_for("sistema.menu_vendas"),
            texto_voltar="← Voltar"
        )

    arquivo = request.files.get("arquivo")

    if not arquivo or arquivo.filename == "":
        flash("Selecione um arquivo.", "error")
        return render_template(
            "vendas_importar_diarias.html",
            nome_empresa=nome_empresa,
            job_id=None
        )

    conn = get_connection()
    cur = conn.cursor()
    tmp = None
    wb = None
    job_id = None

    try:
        job_id = criar_job_importacao(cur, cod_empresa, "vendas_diarias")
        conn.commit()

        atualizar_job_importacao(
            cur,
            job_id,
            etapa="upload",
            mensagem="Recebendo arquivo...",
            percentual=5
        )
        conn.commit()

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        arquivo.save(tmp.name)
        tmp.close()

        atualizar_job_importacao(
            cur,
            job_id,
            etapa="leitura",
            mensagem="Abrindo planilha...",
            percentual=15
        )
        conn.commit()

        wb = load_workbook(tmp.name, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        ws.reset_dimensions()

        cur.execute("""
            SELECT cod_filial, nome_filial_importacao
            FROM filiais
            WHERE cod_empresa = %s
        """, (cod_empresa,))
        filiais_db = cur.fetchall()

        mapa_filiais = {}
        for cod_filial, nome_importacao in filiais_db:
            nome_norm = normalizar_nome_filial_importacao(nome_importacao)

            if nome_norm:
                mapa_filiais[nome_norm] = int(cod_filial)

        # --- validação prévia de filiais ---
        nomes_excel = set()
        for row_val in ws.iter_rows(min_row=1, values_only=True):
            col_a = str(row_val[0] if row_val else "").strip()
            if col_a.upper().startswith("FILIAL:"):
                nome_filial_planilha = col_a.split(":", 1)[1].strip()
                if nome_filial_planilha:
                    nomes_excel.add(nome_filial_planilha)
        ws.reset_dimensions()

        nomes_db_norm = {normalizar_nome_filial_importacao(ni): ni for _, ni in filiais_db if ni}

        filiais_excel_sem_cadastro = [
            n for n in sorted(nomes_excel)
            if localizar_filial_por_nome_importacao(n, mapa_filiais) is None
        ]
        filiais_db_sem_excel = [
            ni for norm_ni, ni in nomes_db_norm.items()
            if not any(
                normalizar_nome_filial_importacao(n)[:len(norm_ni)] == norm_ni
                for n in nomes_excel
            )
        ]

        if filiais_excel_sem_cadastro:
            atualizar_job_importacao(
                cur, job_id, etapa="validacao",
                mensagem="Aviso: filiais no Excel sem cadastro",
                percentual=16
            )
            conn.commit()
            flash(
                "ATENÇÃO — As seguintes filiais estão no Excel mas NÃO foram encontradas no cadastro (não serão importadas): "
                + ", ".join(filiais_excel_sem_cadastro),
                "error"
            )

        if filiais_db_sem_excel:
            flash(
                "ATENÇÃO — As seguintes filiais do cadastro NÃO foram encontradas no arquivo Excel e NÃO serão importadas: "
                + ", ".join(sorted(filiais_db_sem_excel)),
                "error"
            )
        # --- fim validação ---

        dados = []
        linhas_ignoradas = 0
        filial_atual = None
        data_atual = None
        dia_semana_atual = ""
        data_ini = None
        data_fim = None
        linhas_lidas = 0

        for row in ws.iter_rows(min_row=1, values_only=True):
            linhas_lidas += 1

            if linhas_lidas % 500 == 0:
                percentual_leitura = min(55, 20 + (linhas_lidas // 500))
                atualizar_job_importacao(
                    cur,
                    job_id,
                    etapa="leitura",
                    mensagem=f"Lendo planilha... {linhas_lidas} linhas analisadas",
                    percentual=percentual_leitura,
                    linhas_processadas=linhas_lidas
                )
                conn.commit()

            linha = list(row or [])

            while len(linha) < 11:
                linha.append(None)

            col_a = str(linha[0] or "").strip()
            col_b = str(linha[1] or "").strip()
            col_d = str(linha[3] or "").strip()

            if col_a.upper().startswith("FILIAL:"):
                nome_filial_planilha = col_a.split(":", 1)[1].strip()

                filial_atual = localizar_filial_por_nome_importacao(
                    nome_filial_planilha,
                    mapa_filiais
                )

                data_atual = None
                dia_semana_atual = ""
                continue

            texto_a_norm = re.sub(r"\s+", "", col_a.upper())

            if texto_a_norm in ("DATA", "DATA:"):
                texto_data = col_b
                m = re.search(r"(\d{2}/\d{2}/\d{4})", texto_data)

                if m:
                    try:
                        data_atual = datetime.strptime(m.group(1), "%d/%m/%Y").date()

                        parte_dia = texto_data.replace(m.group(1), "").strip()
                        parte_dia = parte_dia.lstrip("-").strip()
                        dia_semana_atual = parte_dia

                        if data_ini is None or data_atual < data_ini:
                            data_ini = data_atual

                        if data_fim is None or data_atual > data_fim:
                            data_fim = data_atual

                    except Exception:
                        data_atual = None
                        dia_semana_atual = ""
                else:
                    data_atual = None
                    dia_semana_atual = ""

                continue

            if filial_atual is None or data_atual is None:
                continue

            if col_d.upper().startswith("SUBTOTAL"):
                continue

            codigo = col_a
            descricao = col_b

            if not codigo or not descricao:
                continue

            if codigo.upper() in ("DATA:", "FILIAL:", "DATA", "FILIAL"):
                continue

            custo = para_float(linha[2])
            preco_venda = para_float(linha[4])
            quantidade = para_float(linha[7])
            valor = para_float(linha[9])
            margem_bruta = para_float(linha[10])

            if quantidade == 0 and valor == 0 and margem_bruta == 0:
                linhas_ignoradas += 1
                continue

            dados.append((
                cod_empresa,
                filial_atual,
                data_atual,
                dia_semana_atual,
                codigo,
                descricao,
                custo,
                preco_venda,
                quantidade,
                valor,
                margem_bruta
            ))

        if not data_ini or not data_fim:
            atualizar_job_importacao(
                cur,
                job_id,
                status="erro",
                etapa="erro",
                mensagem="Não foi possível identificar o período no arquivo.",
                percentual=100
            )
            conn.commit()

            flash("Não foi possível identificar o período no arquivo.", "error")
            return render_template(
                "vendas_importar_diarias.html",
                nome_empresa=nome_empresa,
                job_id=job_id
            )

        if not dados:
            atualizar_job_importacao(
                cur,
                job_id,
                status="erro",
                etapa="erro",
                mensagem="Nenhum dado válido encontrado no arquivo.",
                percentual=100
            )
            conn.commit()

            flash("Nenhum dado válido encontrado no arquivo.", "error")
            return render_template(
                "vendas_importar_diarias.html",
                nome_empresa=nome_empresa,
                job_id=job_id
            )

        atualizar_job_importacao(
            cur,
            job_id,
            etapa="limpeza",
            mensagem="Limpando dados antigos do período...",
            percentual=60,
            total_linhas=len(dados),
            linhas_processadas=0
        )
        conn.commit()

        limpar_importacao_diaria_periodo(cur, conn, cod_empresa, data_ini, data_fim, lote=5000)

        atualizar_job_importacao(
            cur,
            job_id,
            etapa="insercao",
            mensagem="Gravando novos dados...",
            percentual=70,
            total_linhas=len(dados),
            linhas_processadas=0
        )
        conn.commit()

        lote = 5000
        sql_insert = """
            INSERT INTO vendas_diarias (
                cod_empresa,
                cod_filial,
                data,
                dia_semana,
                codigo_produto,
                descricao,
                custo,
                preco_venda,
                quantidade,
                valor,
                margem_bruta
            )
            VALUES %s
        """

        for i in range(0, len(dados), lote):
            bloco = dados[i:i + lote]

            execute_values(
                cur,
                sql_insert,
                bloco,
                page_size=lote
            )

            processados = min(i + len(bloco), len(dados))
            percentual = 70 + int((processados / len(dados)) * 25)

            atualizar_job_importacao(
                cur,
                job_id,
                etapa="insercao",
                mensagem=f"Gravando novos dados... {processados} de {len(dados)}",
                percentual=min(percentual, 95),
                total_linhas=len(dados),
                linhas_processadas=processados
            )

            conn.commit()

        atualizar_job_importacao(
            cur,
            job_id,
            status="concluido",
            etapa="finalizado",
            mensagem="Importação concluída com sucesso.",
            percentual=100,
            total_linhas=len(dados),
            linhas_processadas=len(dados)
        )
        conn.commit()

        msg = (
            f"{len(dados)} registros importados com sucesso. "
            f"Período identificado no arquivo: {data_ini.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}."
        )
        if linhas_ignoradas > 0:
            msg += f" {linhas_ignoradas} linhas foram ignoradas."

        flash(msg, "success")

    except Exception as e:
        conn.rollback()

        try:
            if job_id:
                atualizar_job_importacao(
                    cur,
                    job_id,
                    status="erro",
                    etapa="erro",
                    mensagem=str(e),
                    percentual=100
                )
                conn.commit()
        except Exception:
            pass

        flash(f"Erro: {e}", "error")

    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

        cur.close()
        conn.close()

        if tmp is not None and os.path.exists(tmp.name):
            os.unlink(tmp.name)

    return render_template(
        "vendas_importar_diarias.html",
        nome_empresa=nome_empresa,
        job_id=job_id,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar"
    )

# =========================
# CONSULTAS
# =========================
@vendas_bp.route("/consultas")
@permissao_obrigatoria("VENDAS", "CONSULTAS", redirecionar_para="sistema.menu_vendas")
def vendas_consultas():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    data_ini_padrao, data_fim_padrao = obter_periodo_padrao_consulta()

    data_ini_txt = (request.args.get("data_ini") or "").strip()
    data_fim_txt = (request.args.get("data_fim") or "").strip()
    cod_filial_sel = (request.args.get("cod_filial") or "").strip()

    data_ini = para_data_excel(data_ini_txt) if data_ini_txt else data_ini_padrao
    data_fim = para_data_excel(data_fim_txt) if data_fim_txt else data_fim_padrao

    if not data_ini:
        data_ini = data_ini_padrao
    if not data_fim:
        data_fim = data_fim_padrao

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    linhas_totais = []
    linhas_filial = []
    filial_atual = None
    filial_anterior = None
    filial_proxima = None

    try:
        filiais = obter_filiais_ativas(cur, cod_empresa)

        if cod_filial_sel:
            try:
                cod_filial_int = int(cod_filial_sel)
            except Exception:
                cod_filial_int = None

            if cod_filial_int is not None:
                cur.execute("""
                    SELECT
                        data,
                        dia_semana,
                        SUM(quantidade) AS quantidade,
                        SUM(valor) AS valor,
                        SUM(margem_bruta) AS mb
                    FROM vendas_diarias
                    WHERE cod_empresa = %s
                      AND cod_filial = %s
                      AND data BETWEEN %s AND %s
                    GROUP BY data, dia_semana
                    ORDER BY data ASC
                """, (cod_empresa, cod_filial_int, data_ini, data_fim))
                linhas_filial = cur.fetchall()

                idx = next((i for i, f in enumerate(filiais) if int(f["cod_filial"]) == cod_filial_int), None)
                if idx is not None:
                    filial_atual = filiais[idx]
                    if idx > 0:
                        filial_anterior = filiais[idx - 1]
                    if idx < len(filiais) - 1:
                        filial_proxima = filiais[idx + 1]
        else:
            cur.execute("""
                SELECT
                    data,
                    dia_semana,
                    SUM(quantidade) AS quantidade,
                    SUM(valor) AS valor,
                    SUM(margem_bruta) AS mb
                FROM vendas_diarias
                WHERE cod_empresa = %s
                  AND data BETWEEN %s AND %s
                GROUP BY data, dia_semana
                ORDER BY data ASC
            """, (cod_empresa, data_ini, data_fim))
            linhas_totais = cur.fetchall()

    finally:
        cur.close()
        conn.close()

    for linha in linhas_totais:
        quantidade = float(linha["quantidade"] or 0)
        valor = float(linha["valor"] or 0)
        mb = float(linha["mb"] or 0)

        linha["quantidade"] = quantidade
        linha["valor"] = valor
        linha["mb"] = mb
        linha["mun"] = (mb / quantidade) if quantidade else 0.0
        linha["data_fmt"] = formatar_data_brasil(linha["data"])

    for linha in linhas_filial:
        quantidade = float(linha["quantidade"] or 0)
        valor = float(linha["valor"] or 0)
        mb = float(linha["mb"] or 0)

        linha["quantidade"] = quantidade
        linha["valor"] = valor
        linha["mb"] = mb
        linha["mun"] = (mb / quantidade) if quantidade else 0.0
        linha["data_fmt"] = formatar_data_brasil(linha["data"])

    linhas_totais = aplicar_heatmap_consulta(linhas_totais)
    linhas_filial = aplicar_heatmap_consulta(linhas_filial)

    dados_grafico = serie_grafico_consulta(linhas_filial if cod_filial_sel else linhas_totais)

    return render_template(
        "vendas_consultas.html",
        dados_grafico=dados_grafico,
        nome_empresa=nome_empresa,
        filiais=filiais,
        linhas_totais=linhas_totais,
        linhas_filial=linhas_filial,
        filial_atual=filial_atual,
        filial_anterior=filial_anterior,
        filial_proxima=filial_proxima,
        cod_filial_sel=cod_filial_sel,
        data_ini=data_ini.strftime("%Y-%m-%d"),
        data_fim=data_fim.strftime("%Y-%m-%d"),
        formatar_numero_br=formatar_numero_br,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar"
    )

# --------------------------------------------------------
# IMPORTAR - PROGRESSO
# --------------------------------------------------------

@vendas_bp.route("/diarias/importar/progresso/<job_id>")
@permissao_obrigatoria("VENDAS", "IMPORTAR_VENDAS_DIARIAS", redirecionar_para="sistema.menu_vendas")
def vendas_importar_diarias_progresso(job_id):
    if "cod_empresa" not in session:
        return {"ok": False, "erro": "sessao_expirada"}, 401

    cod_empresa = str(session["cod_empresa"]).strip()

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        cur.execute("""
            SELECT
                job_id,
                cod_empresa,
                tipo_importacao,
                status,
                etapa,
                mensagem,
                percentual,
                total_linhas,
                linhas_processadas
            FROM importacoes_progresso
            WHERE job_id = %s
              AND cod_empresa = %s
        """, (job_id, cod_empresa))

        row = cur.fetchone()

        if not row:
            return {"ok": False, "erro": "job_nao_encontrado"}, 404

        return {
            "ok": True,
            "job": {
                "job_id": row["job_id"],
                "cod_empresa": row["cod_empresa"],
                "tipo_importacao": row["tipo_importacao"],
                "status": row["status"],
                "etapa": row["etapa"],
                "mensagem": row["mensagem"],
                "percentual": row["percentual"],
                "total_linhas": row["total_linhas"],
                "linhas_processadas": row["linhas_processadas"],
            }
        }

    finally:
        cur.close()
        conn.close()

# ------------------------------------------
# CONSULTA POR PRODUTO
# ------------------------------------------

@vendas_bp.route("/consulta_produto")
@permissao_obrigatoria("VENDAS", "CONSULTA_POR_PRODUTO", redirecionar_para="sistema.menu_vendas")
def vendas_consulta_produto():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    hoje = datetime.now().date()
    data_ini_padrao = date(hoje.year, hoje.month, 1)

    cod_filial_sel = (request.args.get("cod_filial") or "").strip()
    data_ini_txt = (request.args.get("data_ini") or "").strip()
    produto_sel = (request.args.get("produto") or "").strip()

    data_ini = para_data_excel(data_ini_txt) if data_ini_txt else data_ini_padrao
    if not data_ini:
        data_ini = data_ini_padrao

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    linhas = []
    filiais = []
    produtos = []

    try:
        cur.execute("""
            SELECT cod_filial, nome_filial
            FROM filiais
            WHERE cod_empresa = %s
              AND ativo = TRUE
            ORDER BY cod_filial
        """, (cod_empresa,))
        filiais = cur.fetchall()

        cur.execute("""
            SELECT DISTINCT descricao
            FROM vendas_diarias
            WHERE cod_empresa = %s
              AND COALESCE(TRIM(descricao), '') <> ''
            ORDER BY descricao
        """, (cod_empresa,))
        produtos = [r["descricao"] for r in cur.fetchall()]

        where = ["cod_empresa = %s", "data >= %s"]
        params = [cod_empresa, data_ini]

        if cod_filial_sel:
            where.append("CAST(cod_filial AS TEXT) = %s")
            params.append(cod_filial_sel)

        if produto_sel:
            where.append("descricao = %s")
            params.append(produto_sel)

        where_sql = " AND ".join(where)

        cur.execute(f"""
            SELECT
                codigo_produto,
                cod_filial,
                data,
                dia_semana,
                descricao,
                custo,
                preco_venda,
                quantidade,
                valor,
                margem_bruta
            FROM vendas_diarias
            WHERE {where_sql}
            ORDER BY data ASC, cod_filial ASC, descricao ASC
        """, params)

        linhas = cur.fetchall()

        mapa_filiais = {int(f["cod_filial"]): f["nome_filial"] for f in filiais}

        for linha in linhas:
            qtd = float(linha["quantidade"] or 0)
            valor = float(linha["valor"] or 0)
            mb = float(linha["margem_bruta"] or 0)
            custo = float(linha["custo"] or 0)
            preco = float(linha["preco_venda"] or 0)

            linha["quantidade"] = qtd
            linha["valor"] = valor
            linha["margem_bruta"] = mb
            linha["custo"] = custo
            linha["preco_venda"] = preco
            linha["mun"] = (mb / qtd) if qtd else 0.0
            linha["data_fmt"] = formatar_data_brasil(linha["data"])
            linha["nome_filial"] = mapa_filiais.get(int(linha["cod_filial"]), str(linha["cod_filial"]))

    finally:
        cur.close()
        conn.close()

    linhas = aplicar_heatmap_variacoes(
        linhas,
        ["custo", "preco_venda", "mun", "quantidade", "valor", "margem_bruta"]
    )

    return render_template(
        "vendas_consulta_produto.html",
        nome_empresa=nome_empresa,
        filiais=filiais,
        produtos=produtos,
        linhas=linhas,
        cod_filial_sel=cod_filial_sel,
        data_ini=data_ini.strftime("%Y-%m-%d"),
        produto_sel=produto_sel,
        formatar_numero_br=formatar_numero_br,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar"
    )


# =========================
# CADASTROS - FILIAIS
# =========================
@vendas_bp.route("/cadastros/filiais")
def vendas_filiais():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))
    return render_template("vendas_filiais.html")

# ------------------------------------------
# MARGEM UNITÁRIA POR DIA
# ------------------------------------------
# Grid produto x posto de um dia só: preço de compra, preço de venda e margem
# unitária, lidos de vendas_diarias.
#
# vendas_diarias traz várias linhas por filial/produto/dia (uma por preço
# praticado), então os três números são médias ponderadas pela quantidade —
# nada é persistido.
#
# A sequência dos produtos é a do cadastro de combustíveis (o mesmo dos
# tanques, em Operações), com os adicionais logo depois.
#
# vendas_diarias guarda a descrição do sistema de origem ("GASOLINA C COMUM",
# "DIESEL COMUM", "DIESEL S10 FROTA."), então cada descrição é encostada no
# produto pela MESMA regra de palavra-chave da Consulta de Estoques
# (services/estoques_service.py) — mudar uma sem a outra faria as duas telas
# divergirem.

BLOCOS_MARGEM_UNITARIA = [
    ("compra", "Compra"),
    ("venda", "Venda"),
    ("margem", "Margem"),
]

# Produtos que aparecem no grid mas NÃO estão em `combustiveis`: são vendidos,
# só não têm tanque. Entram depois dos do cadastro, na ordem daqui.
PRODUTOS_ADICIONAIS_MARGEM = [
    {"cod_produto": "GN", "descricao": "Gás Natural"},
]

# Ordem importa: a primeira palavra encontrada decide. É a ordem do CASE de
# estoques_service, mais o gás natural (antes de GASOL, que "GAS NATURAL" não
# alcança, mas a intenção fica explícita) e o diesel sem sufixo, que no sistema
# de origem vem como "DIESEL COMUM" e é o S500.
# Vendido no posto, mas não é combustível: fica fora do grid e não vale nem
# aviso — a ausência dele não é uma diferença a explicar.
PRODUTOS_IGNORADOS_MARGEM = ("ARLA",)

PALAVRAS_PRODUTO_MARGEM = [
    ("NATURAL", "GN"),
    ("GNV", "GN"),
    ("S10", "C5"),
    ("S500", "C4"),
    ("ADIT", "C2"),
    ("ETAN", "C3"),
    ("PODIUM", "C6"),
    ("GASOL", "C1"),
    ("DIESEL", "C4"),
]


def _cod_produto_das_vendas(descricao_venda):
    """Código do produto para uma descrição vinda de vendas_diarias, pela mesma
    regra de palavra-chave da Consulta de Estoques. Devolve None para o que não
    é combustível (ARLA, por exemplo)."""
    txt = re.sub(r"[^A-Z0-9]", "", str(descricao_venda or "").upper())
    if not txt:
        return None

    if any(palavra in txt for palavra in PRODUTOS_IGNORADOS_MARGEM):
        return None

    for palavra, cod_produto in PALAVRAS_PRODUTO_MARGEM:
        if palavra in txt:
            return cod_produto

    return None


@vendas_bp.route("/margem_unitaria")
@permissao_obrigatoria("VENDAS", "MARGEM_UNITARIA", redirecionar_para="sistema.menu_vendas")
def vendas_margem_unitaria():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    data_txt = (request.args.get("data") or "").strip()
    data_sel = para_data_excel(data_txt) if data_txt else None

    # Os checkboxes só valem quando o formulário foi enviado; na primeira
    # abertura os três vêm marcados.
    if request.args.get("filtrado"):
        mostrar = {
            "compra": request.args.get("mostrar_compra") == "1",
            "venda": request.args.get("mostrar_venda") == "1",
            "margem": request.args.get("mostrar_margem") == "1",
        }
        if not any(mostrar.values()):
            mostrar = {"compra": True, "venda": True, "margem": True}
    else:
        mostrar = {"compra": True, "venda": True, "margem": True}

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    filiais = []
    linhas = []
    sem_cadastro = []

    try:
        if not data_sel:
            cur.execute("""
                SELECT MAX(data) AS ultima
                FROM vendas_diarias
                WHERE cod_empresa = %s
            """, (cod_empresa,))
            row = cur.fetchone() or {}
            hoje = datetime.now(ZoneInfo("America/Recife")).date()
            data_sel = row.get("ultima") or (hoje - timedelta(days=1))

        filiais = obter_filiais_ativas(cur, cod_empresa)

        cur.execute("""
            SELECT cod_produto, descricao
            FROM combustiveis
            WHERE cod_empresa = %s
            ORDER BY cod_produto
        """, (cod_empresa,))
        produtos = list(cur.fetchall() or []) + PRODUTOS_ADICIONAIS_MARGEM

        cur.execute("""
            SELECT
                cod_filial,
                descricao,
                SUM(quantidade) AS quantidade,
                SUM(valor) AS valor,
                SUM(margem_bruta) AS margem_bruta,
                SUM(COALESCE(custo, 0) * quantidade) AS custo_total
            FROM vendas_diarias
            WHERE cod_empresa = %s
              AND data = %s
            GROUP BY cod_filial, descricao
        """, (cod_empresa, data_sel))
        registros = cur.fetchall() or []

    finally:
        cur.close()
        conn.close()

    # (cod_produto, cod_filial) -> somatórios
    acumulado = defaultdict(lambda: {"quantidade": 0.0, "valor": 0.0,
                                     "margem_bruta": 0.0, "custo_total": 0.0})
    nao_casados = set()

    for registro in registros:
        descricao = str(registro["descricao"] or "").strip()
        cod_produto = _cod_produto_das_vendas(descricao)
        if not cod_produto:
            txt = re.sub(r"[^A-Z0-9]", "", descricao.upper())
            if not any(p in txt for p in PRODUTOS_IGNORADOS_MARGEM):
                nao_casados.add(descricao)
            continue

        chave = (cod_produto, int(registro["cod_filial"]))
        acumulado[chave]["quantidade"] += float(registro["quantidade"] or 0)
        acumulado[chave]["valor"] += float(registro["valor"] or 0)
        acumulado[chave]["margem_bruta"] += float(registro["margem_bruta"] or 0)
        acumulado[chave]["custo_total"] += float(registro["custo_total"] or 0)

    sem_cadastro = sorted(nao_casados)
    codigos_filiais = [int(f["cod_filial"]) for f in filiais]

    for produto in produtos:
        cod_produto = str(produto["cod_produto"]).strip()

        # Produto sem venda no dia não vira linha: a empresa que não trabalha
        # Gasolina Especial não precisa ver a faixa vazia dela.
        vendeu = any(
            float(acumulado[(cod_produto, cod_filial)]["quantidade"]) > 0
            for cod_filial in codigos_filiais
            if (cod_produto, cod_filial) in acumulado
        )
        if not vendeu:
            continue

        blocos = []
        for chave, rotulo in BLOCOS_MARGEM_UNITARIA:
            if not mostrar[chave]:
                continue

            valores = []
            for cod_filial in codigos_filiais:
                dados = acumulado.get((cod_produto, cod_filial))
                qtd = float(dados["quantidade"]) if dados else 0.0

                if not dados or qtd <= 0:
                    valores.append(None)
                    continue

                if chave == "compra":
                    valores.append(dados["custo_total"] / qtd)
                elif chave == "venda":
                    valores.append(dados["valor"] / qtd)
                else:
                    valores.append(dados["margem_bruta"] / qtd)

            # Mapa de calor por linha, como no matricial: cada linha se compara
            # com ela mesma, de posto a posto.
            preenchidos = [v for v in valores if v is not None]
            minimo = min(preenchidos) if preenchidos else None
            maximo = max(preenchidos) if preenchidos else None

            celulas = []
            for valor in valores:
                if valor is None or minimo is None:
                    celulas.append({"valor": None, "cor": ""})
                else:
                    # Na compra a escala é invertida: comprar caro é o ruim,
                    # então o preço alto vai para o vermelho.
                    referencia = (minimo + maximo - valor) if chave == "compra" else valor
                    celulas.append({
                        "valor": valor,
                        "cor": cor_excel_51(referencia, minimo, maximo),
                    })

            blocos.append({"chave": chave, "rotulo": rotulo, "celulas": celulas})

        if blocos:
            linhas.append({
                "cod_produto": cod_produto,
                "descricao": produto["descricao"],
                "blocos": blocos,
            })

    return render_template(
        "vendas_margem_unitaria.html",
        nome_empresa=nome_empresa,
        filiais=filiais,
        linhas=linhas,
        mostrar=mostrar,
        sem_cadastro=sem_cadastro,
        data_sel=data_sel.strftime("%Y-%m-%d"),
        formatar_numero_br=formatar_numero_br,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar",
    )


# ------------------------------------------
# VENDAS POR DIA
# ------------------------------------------
# Mesmo grid da Margem Unitária por Dia (produto x posto, um dia só), mas com
# os volumes em vez dos unitários: litros vendidos, dinheiro vendido e margem
# bruta em dinheiro. Tudo somado de vendas_diarias, que traz uma linha por
# preço praticado — aqui não há média ponderada, é soma direta.
#
# A regra de produto, os adicionais e os ignorados são os MESMOS da margem
# unitária (e, por ela, os da Consulta de Estoques): as duas telas não podem
# enxergar produtos diferentes no mesmo dia.

BLOCOS_VENDAS_POR_DIA = [
    ("quantidade", "Quantidade"),
    ("valor", "Valor"),
    ("margem", "Margem Bruta"),
]


@vendas_bp.route("/por_dia")
@permissao_obrigatoria("VENDAS", "VENDAS_POR_DIA", redirecionar_para="sistema.menu_vendas")
def vendas_por_dia():
    if "cod_empresa" not in session:
        return redirect(url_for("auth.index"))

    cod_empresa = str(session["cod_empresa"]).strip()
    nome_empresa = session.get("nome_empresa", "")

    data_txt = (request.args.get("data") or "").strip()
    data_sel = para_data_excel(data_txt) if data_txt else None

    # Os checkboxes só valem quando o formulário foi enviado; na primeira
    # abertura os três vêm marcados.
    if request.args.get("filtrado"):
        mostrar = {
            "quantidade": request.args.get("mostrar_quantidade") == "1",
            "valor": request.args.get("mostrar_valor") == "1",
            "margem": request.args.get("mostrar_margem") == "1",
        }
        if not any(mostrar.values()):
            mostrar = {"quantidade": True, "valor": True, "margem": True}
    else:
        mostrar = {"quantidade": True, "valor": True, "margem": True}

    conn = get_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    filiais = []
    linhas = []
    sem_cadastro = []

    try:
        if not data_sel:
            cur.execute("""
                SELECT MAX(data) AS ultima
                FROM vendas_diarias
                WHERE cod_empresa = %s
            """, (cod_empresa,))
            row = cur.fetchone() or {}
            hoje = datetime.now(ZoneInfo("America/Recife")).date()
            data_sel = row.get("ultima") or (hoje - timedelta(days=1))

        filiais = obter_filiais_ativas(cur, cod_empresa)

        cur.execute("""
            SELECT cod_produto, descricao
            FROM combustiveis
            WHERE cod_empresa = %s
            ORDER BY cod_produto
        """, (cod_empresa,))
        produtos = list(cur.fetchall() or []) + PRODUTOS_ADICIONAIS_MARGEM

        cur.execute("""
            SELECT
                cod_filial,
                descricao,
                SUM(quantidade) AS quantidade,
                SUM(valor) AS valor,
                SUM(margem_bruta) AS margem_bruta
            FROM vendas_diarias
            WHERE cod_empresa = %s
              AND data = %s
            GROUP BY cod_filial, descricao
        """, (cod_empresa, data_sel))
        registros = cur.fetchall() or []

    finally:
        cur.close()
        conn.close()

    # (cod_produto, cod_filial) -> somatórios
    acumulado = defaultdict(lambda: {"quantidade": 0.0, "valor": 0.0,
                                     "margem_bruta": 0.0})
    nao_casados = set()

    for registro in registros:
        descricao = str(registro["descricao"] or "").strip()
        cod_produto = _cod_produto_das_vendas(descricao)
        if not cod_produto:
            txt = re.sub(r"[^A-Z0-9]", "", descricao.upper())
            if not any(p in txt for p in PRODUTOS_IGNORADOS_MARGEM):
                nao_casados.add(descricao)
            continue

        chave = (cod_produto, int(registro["cod_filial"]))
        acumulado[chave]["quantidade"] += float(registro["quantidade"] or 0)
        acumulado[chave]["valor"] += float(registro["valor"] or 0)
        acumulado[chave]["margem_bruta"] += float(registro["margem_bruta"] or 0)

    sem_cadastro = sorted(nao_casados)
    codigos_filiais = [int(f["cod_filial"]) for f in filiais]

    for produto in produtos:
        cod_produto = str(produto["cod_produto"]).strip()

        # Produto sem venda no dia não vira linha, como na margem unitária.
        vendeu = any(
            float(acumulado[(cod_produto, cod_filial)]["quantidade"]) > 0
            for cod_filial in codigos_filiais
            if (cod_produto, cod_filial) in acumulado
        )
        if not vendeu:
            continue

        blocos = []
        for chave, rotulo in BLOCOS_VENDAS_POR_DIA:
            if not mostrar[chave]:
                continue

            valores = []
            for cod_filial in codigos_filiais:
                dados = acumulado.get((cod_produto, cod_filial))
                qtd = float(dados["quantidade"]) if dados else 0.0

                if not dados or qtd <= 0:
                    valores.append(None)
                    continue

                if chave == "quantidade":
                    valores.append(dados["quantidade"])
                elif chave == "valor":
                    valores.append(dados["valor"])
                else:
                    valores.append(dados["margem_bruta"])

            # Mapa de calor por linha: cada linha se compara de posto a posto.
            # Nos três blocos o alto é o bom (vendeu mais, faturou mais, ganhou
            # mais), então não há escala invertida como no preço de compra.
            preenchidos = [v for v in valores if v is not None]
            minimo = min(preenchidos) if preenchidos else None
            maximo = max(preenchidos) if preenchidos else None

            celulas = []
            for valor in valores:
                if valor is None or minimo is None:
                    celulas.append({"valor": None, "cor": ""})
                else:
                    celulas.append({
                        "valor": valor,
                        "cor": cor_excel_51(valor, minimo, maximo),
                    })

            blocos.append({"chave": chave, "rotulo": rotulo, "celulas": celulas})

        if blocos:
            linhas.append({
                "cod_produto": cod_produto,
                "descricao": produto["descricao"],
                "blocos": blocos,
            })

    # Rodapé: um bloco por opção marcada em "Mostrar", somando TODOS os
    # combustíveis do grid (os mesmos que viraram linha) posto a posto. Sai da
    # mesma estrutura das linhas, então esconder um bloco esconde o total dele.
    campos_total = {"quantidade": "quantidade", "valor": "valor",
                    "margem": "margem_bruta"}
    codigos_no_grid = [linha["cod_produto"] for linha in linhas]
    totais = []

    for chave, rotulo in BLOCOS_VENDAS_POR_DIA:
        if not mostrar[chave]:
            continue

        campo = campos_total[chave]
        valores = []
        for cod_filial in codigos_filiais:
            soma = 0.0
            teve = False
            for cod_produto in codigos_no_grid:
                dados = acumulado.get((cod_produto, cod_filial))
                if not dados or float(dados["quantidade"]) <= 0:
                    continue
                soma += float(dados[campo])
                teve = True
            valores.append(soma if teve else None)

        preenchidos = [v for v in valores if v is not None]
        minimo = min(preenchidos) if preenchidos else None
        maximo = max(preenchidos) if preenchidos else None

        celulas = []
        for valor in valores:
            if valor is None or minimo is None:
                celulas.append({"valor": None, "cor": ""})
            else:
                celulas.append({
                    "valor": valor,
                    "cor": cor_excel_51(valor, minimo, maximo),
                })

        totais.append({"chave": chave, "rotulo": rotulo, "celulas": celulas})

    return render_template(
        "vendas_por_dia.html",
        nome_empresa=nome_empresa,
        filiais=filiais,
        linhas=linhas,
        totais=totais,
        mostrar=mostrar,
        sem_cadastro=sem_cadastro,
        data_sel=data_sel.strftime("%Y-%m-%d"),
        formatar_numero_br=formatar_numero_br,
        url_voltar=url_for("sistema.menu_vendas"),
        texto_voltar="← Voltar",
    )
